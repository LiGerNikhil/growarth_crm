from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime, timedelta

from .models import Employee, Attendance, Break
from .forms import DocumentForm


class AttendanceReportView(LoginRequiredMixin, ListView):
    """Attendance Summary Dashboard for Admin and HR"""
    model = Attendance
    template_name = 'accounts/attendance_report.html'
    context_object_name = 'attendances'
    paginate_by = 50
    
    def dispatch(self, request, *args, **kwargs):
        # Check if user has permission (Admin or HR)
        if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):
            messages.error(request, 'You do not have permission to access attendance reports.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        queryset = Attendance.objects.select_related('employee', 'employee__team_leader', 'corrected_by')
        
        # Date range filter
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(date__gte=start_date)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(date__lte=end_date)
            except ValueError:
                pass
        
        # Team Leader filter
        team_leader_id = self.request.GET.get('team_leader')
        if team_leader_id:
            queryset = queryset.filter(employee__team_leader_id=team_leader_id)
        
        # Status filter
        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(final_status=status_filter)
        
        # Employee filter
        employee_id = self.request.GET.get('employee')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.order_by('-date', 'employee__first_name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter values
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['selected_team_leader'] = self.request.GET.get('team_leader', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_employee'] = self.request.GET.get('employee', '')
        
        # Get team leaders for filter dropdown
        context['team_leaders'] = Employee.objects.filter(
            role='team_leader', 
            employment_status='active'
        ).order_by('first_name', 'last_name')
        
        # Get employees for filter dropdown
        context['employees'] = Employee.objects.filter(
            employment_status='active'
        ).order_by('first_name', 'last_name')
        
        # Status choices for filter
        context['status_choices'] = Attendance.STATUS_CHOICES
        
        # Calculate summary statistics from the original queryset (before pagination)
        # We need to get the full queryset without pagination for statistics
        full_queryset = self.get_queryset()
        context['total_records'] = full_queryset.count()
        context['full_days'] = full_queryset.filter(final_status='full_day_present').count()
        context['half_days'] = full_queryset.filter(final_status='half_day').count()
        context['absent_days'] = full_queryset.filter(final_status='absent').count()
        context['manually_corrected'] = full_queryset.filter(correction_status='manually_corrected').count()
        
        # Add calculated data to each attendance record (paginated queryset)
        for attendance in context['attendances']:
            attendance.effective_hours = attendance.effective_work_hours
            # calculated_status is a property, no need to set it
            attendance.needs_buffer_review = self._needs_buffer_review(attendance)
        
        return context
    
    def _needs_buffer_review(self, attendance):
        """Check if attendance needs buffer review (within 15 minutes of full day)"""
        if attendance.correction_status == 'manually_corrected':
            return False
        
        effective_hours = attendance.effective_work_hours
        required_full_day = 8
        buffer = 15 / 60  # 15 minutes
        
        return (effective_hours >= (required_full_day - buffer) and 
                effective_hours < required_full_day)


@require_POST
@login_required
def override_attendance_status(request, attendance_id):
    """Override attendance status manually"""
    print(f"Override request received for attendance {attendance_id}")
    print(f"User: {request.user.employee.full_name}")
    print(f"Request method: {request.method}")
    print(f"POST data: {dict(request.POST)}")
    
    # Check if this is an AJAX request or form submission
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Check permissions
    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):
        print("Permission denied")
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        else:
            messages.error(request, 'Permission denied')
            return redirect('accounts:attendance_report')
    
    try:
        attendance = Attendance.objects.get(id=attendance_id)
        print(f"Found attendance: {attendance.employee.full_name} on {attendance.date}")
    except Attendance.DoesNotExist:
        print("Attendance not found")
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Attendance not found'})
        else:
            messages.error(request, 'Attendance not found')
            return redirect('accounts:attendance_report')
    
    new_status = request.POST.get('status')
    notes = request.POST.get('notes', '')
    
    print(f"New status: {new_status}")
    print(f"Notes: {notes}")
    
    if new_status not in [choice[0] for choice in Attendance.STATUS_CHOICES]:
        print("Invalid status")
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Invalid status'})
        else:
            messages.error(request, 'Invalid status')
            return redirect('accounts:attendance_report')
    
    try:
        # Direct update instead of using manual_correction method
        attendance.final_status = new_status
        attendance.correction_status = 'manually_corrected'
        attendance.corrected_by = request.user.employee
        attendance.corrected_at = timezone.now()
        attendance.correction_notes = notes
        attendance.save()
        
        print("Status updated successfully")
        
        if is_ajax:
            return JsonResponse({'success': True, 'message': 'Attendance status updated successfully!'})
        else:
            messages.success(request, f'Attendance status for {attendance.employee.full_name} on {attendance.date} has been updated.')
            return redirect('accounts:attendance_report')
            
    except Exception as e:
        print(f"Exception occurred: {str(e)}")
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)})
        else:
            messages.error(request, f'Error updating attendance status: {str(e)}')
            return redirect('accounts:attendance_report')


def export_attendance_report_excel(request):
    """Export attendance report to Excel in clean employee-wise format"""
    # Check permissions
    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):
        messages.error(request, 'You do not have permission to export attendance reports.')
        return redirect('accounts:attendance_report')
    
    # Get filtered attendances (same logic as AttendanceReportView)
    queryset = Attendance.objects.select_related('employee', 'employee__team_leader', 'corrected_by')
    
    # Apply same filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    team_leader_id = request.GET.get('team_leader')
    status_filter = request.GET.get('status')
    employee_id = request.GET.get('employee')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(date__lte=end_date)
        except ValueError:
            pass
    
    if team_leader_id:
        queryset = queryset.filter(employee__team_leader_id=team_leader_id)
    
    if status_filter:
        queryset = queryset.filter(final_status=status_filter)
    
    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)
    
    # Order by employee name first, then by date
    queryset = queryset.order_by('employee__first_name', 'employee__last_name', 'date')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    employee_header_font = Font(bold=True, color="FFFFFF", size=14)
    employee_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    date_font = Font(bold=True, size=11)
    date_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Main headers
    main_headers = [
        'Employee Name', 'Team Leader', 'Position', 'Date', 'Start Time', 'End Time', 
        'Break Hours', 'Effective Hours', 'Status', 'Corrected By'
    ]
    
    # Write main headers
    for col_num, header in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Group data by employee
    employees_data = {}
    for attendance in queryset:
        emp_id = attendance.employee.id
        if emp_id not in employees_data:
            employees_data[emp_id] = {
                'employee': attendance.employee,
                'attendances': []
            }
        employees_data[emp_id]['attendances'].append(attendance)
    
    # Write data row by row
    current_row = 2
    
    for emp_id, emp_data in employees_data.items():
        employee = emp_data['employee']
        attendances = emp_data['attendances']
        
        # Write employee info for each attendance record
        for attendance in attendances:
            # Employee Name
            ws.cell(row=current_row, column=1, value=employee.full_name)
            
            # Team Leader
            ws.cell(row=current_row, column=2, 
                   value=employee.team_leader.full_name if employee.team_leader else 'N/A')
            
            # Position
            ws.cell(row=current_row, column=3, value=employee.position or 'N/A')
            
            # Date
            ws.cell(row=current_row, column=4, value=attendance.date.strftime('%d-%m-%Y'))
            
            # Start Time
            ws.cell(row=current_row, column=5, 
                   value=attendance.start_time.strftime('%H:%M') if attendance.start_time else 'N/A')
            
            # End Time
            ws.cell(row=current_row, column=6, 
                   value=attendance.end_time.strftime('%H:%M') if attendance.end_time else 'N/A')
            
            # Break Hours
            ws.cell(row=current_row, column=7, value=f"{attendance.break_hours or 0:.1f}h")
            
            # Effective Hours
            ws.cell(row=current_row, column=8, value=f"{attendance.effective_work_hours:.1f}h")
            
            # Status with color coding
            status_cell = ws.cell(row=current_row, column=9, value=attendance.get_final_status_display())
            if attendance.final_status == 'full_day_present':
                status_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif attendance.final_status == 'half_day':
                status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif attendance.final_status == 'absent':
                status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            # Correction info
            if attendance.correction_status == 'manually_corrected':
                corrected_by = attendance.corrected_by.full_name if attendance.corrected_by else 'N/A'
                ws.cell(row=current_row, column=10, value=f"Corrected by {corrected_by}")
            else:
                ws.cell(row=current_row, column=10, value="Original")
            
            current_row += 1
        
        # Add a blank row between employees for better readability
        current_row += 1
    
    # Add summary section at the bottom
    summary_row = current_row + 2
    
    # Summary headers
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = employee_header_font
    ws.cell(row=summary_row, column=1).fill = employee_header_fill
    
    summary_row += 1
    
    # Calculate summary statistics
    total_records = queryset.count()
    full_days = queryset.filter(final_status='full_day_present').count()
    half_days = queryset.filter(final_status='half_day').count()
    absent_days = queryset.filter(final_status='absent').count()
    manually_corrected = queryset.filter(correction_status='manually_corrected').count()
    
    # Summary data
    ws.cell(row=summary_row, column=1, value="Total Records:")
    ws.cell(row=summary_row, column=2, value=total_records)
    
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="Full Days:")
    ws.cell(row=summary_row, column=2, value=full_days)
    
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="Half Days:")
    ws.cell(row=summary_row, column=2, value=half_days)
    
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="Absent Days:")
    ws.cell(row=summary_row, column=2, value=absent_days)
    
    summary_row += 1
    ws.cell(row=summary_row, column=1, value="Manually Corrected:")
    ws.cell(row=summary_row, column=2, value=manually_corrected)
    
    # Adjust column widths for better readability
    column_widths = [25, 20, 15, 12, 10, 10, 12, 12, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Add borders to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to data range
    for row in ws.iter_rows(min_row=2, max_row=current_row-1, min_col=1, max_col=10):
        for cell in row:
            if cell.value:  # Only apply border to cells with data
                cell.border = thin_border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response
