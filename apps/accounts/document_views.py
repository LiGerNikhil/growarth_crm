from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from .models import Employee, Document
from .forms import DocumentForm


class DocumentEmployeeListView(LoginRequiredMixin, ListView):
    """List of all employees for document management"""
    model = Employee
    template_name = 'accounts/document_employee_list.html'
    context_object_name = 'employees'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Employee.objects.all().order_by('first_name', 'last_name')
        search_query = self.request.GET.get('search', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(employee_id__icontains=search_query) |
                Q(position__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class DocumentListView(LoginRequiredMixin, ListView):
    """List of documents for a specific employee"""
    model = Document
    template_name = 'accounts/document_list.html'
    context_object_name = 'documents'
    paginate_by = 20
    
    def get_queryset(self):
        employee_id = self.kwargs.get('employee_id')
        self.employee = get_object_or_404(Employee, id=employee_id)
        return Document.objects.filter(employee=self.employee).order_by('-uploaded_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employee'] = self.employee
        context['has_bank_details'] = self.employee.documents.filter(document_type='bank_details').exists()
        
        # Add file names for documents
        for document in context['documents']:
            if document.file:
                document.filename = document.file.name.split('/')[-1]  # Get just the filename
            else:
                document.filename = None
                
        return context


class DocumentCreateView(LoginRequiredMixin, CreateView):
    """Create a new document for an employee"""
    model = Document
    form_class = DocumentForm
    template_name = 'accounts/document_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.kwargs.get('employee_id')
        context['employee'] = get_object_or_404(Employee, id=employee_id)
        return context
    
    def form_valid(self, form):
        employee_id = self.kwargs.get('employee_id')
        employee = get_object_or_404(Employee, id=employee_id)
        form.instance.employee = employee
        form.instance.uploaded_by = self.request.user
        messages.success(self.request, f'Document "{form.instance.title}" uploaded successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        employee_id = self.kwargs.get('employee_id')
        return reverse_lazy('accounts:document_list', kwargs={'employee_id': employee_id})

       
class DocumentDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a document"""
    model = Document
    template_name = 'accounts/document_confirm_delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add filename for display
        if self.object.file:
            context['filename'] = self.object.file.name.split('/')[-1]
        else:
            context['filename'] = None
        return context
    
    def get_success_url(self):
        return reverse_lazy('accounts:document_list', kwargs={'employee_id': self.object.employee.id})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, f'Document "{self.object.title}" deleted successfully!')
        return super().delete(request, *args, **kwargs)


def export_bank_details_excel(request):
    """Export bank details to Excel file"""
    # Get all bank detail documents
    bank_documents = Document.objects.filter(
        document_type='bank_details'
    ).select_related('employee').order_by('employee__first_name', 'employee__last_name')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Details"
    
    # Define headers
    headers = [
        'Employee Name', 'Employee ID', 'Email', 'Department', 'Position',
        'Bank Name', 'Account Number', 'IFSC Code', 'Branch Name', 'Document Title'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    row_num = 2
    for doc in bank_documents:
        employee = doc.employee
        
        ws.cell(row=row_num, column=1, value=employee.full_name)
        ws.cell(row=row_num, column=2, value=employee.employee_id or '')
        ws.cell(row=row_num, column=3, value=employee.email)
        ws.cell(row=row_num, column=4, value=employee.get_department_display() or '')
        ws.cell(row=row_num, column=5, value=employee.position or '')
        ws.cell(row=row_num, column=6, value=doc.bank_name or '')
        ws.cell(row=row_num, column=7, value=doc.account_number or '')
        ws.cell(row=row_num, column=8, value=doc.ifsc_code or '')
        ws.cell(row=row_num, column=9, value=doc.branch_name or '')
        ws.cell(row=row_num, column=10, value=doc.title)
        
        row_num += 1
    
    # Adjust column widths
    column_widths = [20, 15, 25, 15, 20, 20, 18, 15, 20, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bank_details_export.xlsx"'
    
    return response


def download_document(request, pk):
    """Download a document file"""
    document = get_object_or_404(Document, id=pk)
    
    if document.file:
        response = HttpResponse(document.file.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{document.file.name.split("/")[-1]}"'
        return response
    
    messages.error(request, 'File not found.')
    return redirect('accounts:document_list', employee_id=document.employee.id)
