import logging

from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth import authenticate, login, logout

from django.contrib.auth.views import LogoutView

from django.contrib.auth.decorators import login_required

from django.contrib.auth.models import User

from django.contrib import messages

from django.http import JsonResponse, HttpResponse

from django.urls import reverse_lazy

from django.db import models, transaction

from .models import Employee, GeolocationLoginAttempt

from django.db.models import Sum, Count, Q

from django.core.paginator import Paginator

from django.utils import timezone

from django.views.decorators.http import require_POST, require_http_methods

from .decorators import team_leader_required, manager_required, admin_required, mis_access_required

from .forms import MISReportForm, MISFilterForm

from .models import MISReport

import openpyxl

from openpyxl.styles import Font, Alignment, PatternFill

from datetime import datetime, date

from django.views.generic import ListView, DetailView, CreateView, UpdateView

from django.db.models import Q

from django.contrib.admin.views.decorators import staff_member_required

from django.utils import timezone

from threading import local

# Initialize logger
logger = logging.getLogger(__name__)

from datetime import datetime



from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger



from .models import Employee, OnboardingTask, Document, ManagerRequest, Team, ExcelBatch, ExcelBatchRow, Lead, LeadForward, LoginStatistics, UserActivityLog, LeadActivityLog, Attendance, Break, Violation, Leave, LeaveApproval

# Import CRM Lead model for Google Sheets integration
from apps.crm.models import Lead as CrmLead

from .forms import (

    StaffForm, OnboardingTaskForm, DocumentUploadForm, EmployeeSearchForm, ManagerRequestForm, TeamForm, ExcelUploadForm, ExcelBatchProcessForm, LoginStatisticsForm, LoginStatisticsFilterForm,

    AttendanceForm, AttendanceFilterForm, ViolationForm, AttendanceStatsForm, LeaveApplicationForm, LeaveApprovalForm, LeaveFilterForm

)

from .permissions import role_required, can_manage_user, can_view_user, RoleRequiredMixin, UserManagementMixin





# Middleware to set current user in thread local for activity tracking

class ThreadLocalMiddleware:

    def __init__(self, get_response):

        self.get_response = get_response



    def __call__(self, request):

        local.user = request.user if request.user.is_authenticated else None

        response = self.get_response(request)

        return response





def login_view(request):
    """
    Login view for user authentication
    """
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}! You have successfully logged in.')
            
            # Check for next parameter
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('core:dashboard')
        else:
            messages.error(request, 'Invalid username or password. Please try again.')
    
    return render(request, 'accounts/login.html')


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def geolocation_logs(request):
    """View for admins to see geolocation login attempts"""
    # Check if user has permission (Admin or HR)
    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):
        messages.error(request, 'You do not have permission to access geolocation logs.')
        return redirect('core:dashboard')
    
    # Get filter parameters
    days = int(request.GET.get('days', 30))
    attempt_type = request.GET.get('attempt_type', '')
    user_filter = request.GET.get('user', '')
    
    # Get base queryset
    attempts = GeolocationLoginAttempt.objects.select_related('user', 'employee')
    
    # Filter by date range
    from django.utils import timezone
    cutoff_date = timezone.now() - timezone.timedelta(days=days)
    attempts = attempts.filter(created_at__gte=cutoff_date)
    
    # Apply filters
    if attempt_type:
        attempts = attempts.filter(attempt_type=attempt_type)
    
    if user_filter:
        attempts = attempts.filter(
            models.Q(user__username__icontains=user_filter) |
            models.Q(user__first_name__icontains=user_filter) |
            models.Q(user__last_name__icontains=user_filter)
        )
    
    # Order by latest first
    attempts = attempts.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(attempts, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'days': days,
        'selected_attempt_type': attempt_type,
        'user_filter': user_filter,
        'attempt_types': GeolocationLoginAttempt.ATTEMPT_TYPES,
        'total_attempts': attempts.count(),
        'blocked_attempts': attempts.filter(success=False, bypass_allowed=False).count(),
        'success_attempts': attempts.filter(success=True).count(),
    }
    
    return render(request, 'accounts/geolocation_logs.html', context)


@login_required
def export_geolocation_logs(request):
    """Export geolocation logs to Excel"""
    # Check permissions
    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):
        messages.error(request, 'You do not have permission to export geolocation logs.')
        return redirect('core:dashboard')
    
    # Get filter parameters (same as geolocation_logs view)
    days = int(request.GET.get('days', 30))
    attempt_type = request.GET.get('attempt_type', '')
    user_filter = request.GET.get('user', '')
    
    # Get base queryset
    attempts = GeolocationLoginAttempt.objects.select_related('user', 'employee')
    
    # Filter by date range
    from django.utils import timezone
    cutoff_date = timezone.now() - timezone.timedelta(days=days)
    attempts = attempts.filter(created_at__gte=cutoff_date)
    
    # Apply filters
    if attempt_type:
        attempts = attempts.filter(attempt_type=attempt_type)
    
    if user_filter:
        attempts = attempts.filter(
            models.Q(user__username__icontains=user_filter) |
            models.Q(user__first_name__icontains=user_filter) |
            models.Q(user__last_name__icontains=user_filter)
        )
    
    # Order by latest first
    attempts = attempts.order_by('-created_at')
    
    # Create workbook and worksheet
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geolocation Login Logs"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Main headers
    main_headers = [
        'Employee Name', 'Username', 'Role', 'Attempt Type', 'Status', 
        'Latitude', 'Longitude', 'Distance from Office', 'IP Address', 
        'User Agent', 'Error Message', 'Created At (IST)', 'Bypass Allowed'
    ]
    
    # Write main headers
    for col_num, header in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    row_num = 2
    for attempt in attempts:
        # Employee Name
        ws.cell(row=row_num, column=1, value=attempt.employee_name)
        
        # Username
        ws.cell(row=row_num, column=2, value=attempt.user.username)
        
        # Role
        ws.cell(row=row_num, column=3, value=attempt.get_user_role_display() or 'N/A')
        
        # Attempt Type
        ws.cell(row=row_num, column=4, value=attempt.get_attempt_type_display())
        
        # Status
        status_cell = ws.cell(row=row_num, column=5, value='Success' if attempt.success else 'Failed')
        if attempt.success:
            status_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Latitude
        ws.cell(row=row_num, column=6, value=attempt.user_latitude or 'N/A')
        
        # Longitude
        ws.cell(row=row_num, column=7, value=attempt.user_longitude or 'N/A')
        
        # Distance from Office
        ws.cell(row=row_num, column=8, value=attempt.formatted_distance)
        
        # IP Address
        ws.cell(row=row_num, column=9, value=attempt.ip_address or 'N/A')
        
        # User Agent (truncated)
        user_agent = attempt.user_agent or 'N/A'
        if len(user_agent) > 50:
            user_agent = user_agent[:50] + '...'
        ws.cell(row=row_num, column=10, value=user_agent)
        
        # Error Message
        error_msg = attempt.error_message or ''
        if len(error_msg) > 100:
            error_msg = error_msg[:100] + '...'
        ws.cell(row=row_num, column=11, value=error_msg)
        
        # Created At (IST)
        ws.cell(row=row_num, column=12, value=attempt.get_created_at_ist())
        
        # Bypass Allowed
        ws.cell(row=row_num, column=13, value='Yes' if attempt.bypass_allowed else 'No')
        
        row_num += 1
    
    # Adjust column widths
    column_widths = [25, 20, 15, 20, 10, 15, 15, 18, 15, 30, 25, 20, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Add borders to all data cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to data range
    for row in ws.iter_rows(min_row=2, max_row=row_num-1, min_col=1, max_col=13):
        for cell in row:
            if cell.value:
                cell.border = thin_border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="geolocation_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response





class CustomLogoutView(LogoutView):

    """

    Custom logout view with success message

    """

    next_page = reverse_lazy('accounts:logout_success')

    

    def dispatch(self, request, *args, **kwargs):

        messages.success(request, 'You have been successfully logged out.')

        return super().dispatch(request, *args, **kwargs)





def logout_success(request):

    """

    Custom logout success page

    """

    return render(request, 'accounts/logout.html')





@login_required

def profile_view(request):

    """

    User profile view

    """

    try:

        employee = request.user.employee

    except Employee.DoesNotExist:

        employee = None

    

    return render(request, 'accounts/profile.html', {'employee': employee})





@login_required

def settings_view(request):

    """

    Account settings view

    """

    return render(request, 'accounts/settings.html')





# Employee Management Views

@login_required

def lead_forward_dashboard(request):
    """
    Dashboard for admins to forward Google Sheets leads to employees
    """
    if not request.user.employee.is_admin and not request.user.employee.is_superadmin and not request.user.employee.is_manager:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('core:dashboard')
    
    # Get Google Sheets leads that haven't been assigned yet
    google_leads = CrmLead.objects.filter(
        synced_with_google=True
    ).filter(
        Q(assigned_to__isnull=True) | Q(assigned_to=None)
    ).order_by('-created_at')
    
    # Debug: Print query and count
    print(f"Google leads query: {google_leads.query}")
    print(f"Google leads count: {google_leads.count()}")
    print(f"Total CrmLead count: {CrmLead.objects.count()}")
    print(f"Synced with Google count: {CrmLead.objects.filter(synced_with_google=True).count()}")
    
    # Get forwarded leads tracking
    forwarded_leads = LeadForward.objects.select_related('forwarded_to', 'forwarded_by').order_by('-forwarded_at')
    
    # Get employees for forwarding dropdown
    employees = Employee.objects.filter(
        employment_status='active'
    ).exclude(
        role__in=['admin', 'superadmin']
    ).order_by('first_name', 'last_name')
    
    context = {
        'google_leads': google_leads,
        'forwarded_leads': forwarded_leads,
        'employees': employees,
        'title': 'Google Sheets Lead Forwarding Dashboard',
    }
    
    return render(request, 'accounts/lead_forward_dashboard.html', context)


def forward_lead(request, lead_id):
    """
    Forward a specific Google Sheets lead to an employee
    """
    if not request.user.employee.is_admin and not request.user.employee.is_superadmin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            lead = CrmLead.objects.get(id=lead_id)
            employee_id = request.POST.get('employee_id')
            message = request.POST.get('message', '')
            priority = request.POST.get('priority', 'medium')
            
            employee = Employee.objects.get(id=employee_id)
            
            # Check if already forwarded
            if LeadForward.objects.filter(
                lead_id=lead.id,
                forwarded_to=employee
            ).exists():
                return JsonResponse({'error': 'Lead already forwarded to this employee'}, status=400)
            
            # Create forward record with lead data stored directly
            forward = LeadForward.objects.create(
                lead_id=str(lead.id),
                lead_name=lead.name,
                lead_email=lead.email,
                lead_phone=lead.phone or '',
                lead_source=lead.get_source_display(),
                forwarded_to=employee,
                forwarded_by=request.user.employee,
                notes=message,
                status='forwarded'
            )
            
            # Update lead to mark as assigned
            lead.assigned_to = request.user
            lead.save()
            
            # Log activity
            from apps.crm.models import LeadActivity
            LeadActivity.objects.create(
                lead=lead,
                activity_type='note',
                description=f"Lead forwarded to {employee.get_full_name()}. Priority: {priority}. Message: {message}",
                created_by=request.user
            )
            
            return JsonResponse({'success': True, 'message': 'Lead forwarded successfully'})
            
        except CrmLead.DoesNotExist:
            return JsonResponse({'error': 'Lead not found'}, status=404)
        except Employee.DoesNotExist:
            return JsonResponse({'error': 'Employee not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def my_forwarded_leads(request):
    """
    View for employees to see Google Sheets leads forwarded to them
    """
    try:
        current_employee = request.user.employee
        
        # Get leads forwarded to current employee
        forwarded_leads = LeadForward.objects.filter(
            forwarded_to=current_employee
        ).select_related('forwarded_by').order_by('-forwarded_at')
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            forwarded_leads = forwarded_leads.filter(status=status_filter)
        
        context = {
            'forwarded_leads': forwarded_leads,
            'current_employee': current_employee,
            'title': 'My Forwarded Google Leads',
            'status_filter': status_filter,
        }
        
        return render(request, 'accounts/my_forwarded_leads.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'You need an employee profile to access this page.')
        return redirect('core:dashboard')


def update_forward_status(request, forward_id):
    """
    Update the status of a forwarded Google Sheets lead
    """
    if request.method == 'POST':
        try:
            current_employee = request.user.employee
            forward = LeadForward.objects.get(
                id=forward_id,
                forwarded_to=current_employee
            )
            
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            if new_status in ['accepted', 'rejected', 'completed']:
                forward.status = new_status
                
                if new_status == 'accepted' and not forward.accepted_at:
                    from django.utils import timezone
                    forward.accepted_at = timezone.now()
                
                if new_status == 'completed':
                    from django.utils import timezone
                    forward.completed_at = timezone.now()
                
                forward.save()
                
                # Log activity only if we can find the original lead
                from apps.crm.models import LeadActivity, Lead as CrmLead
                try:
                    if forward.lead_id and forward.lead_id.isdigit():
                        original_lead = CrmLead.objects.get(id=int(forward.lead_id))
                        
                        LeadActivity.objects.create(
                            lead=original_lead,
                            activity_type='status_change',
                            description=f"Forward status changed to {new_status}. {notes}",
                            created_by=request.user
                        )
                        
                        # Update original lead status if possible
                        if new_status == 'accepted':
                            original_lead.status = 'contacted'
                            original_lead.save()
                except (CrmLead.DoesNotExist, ValueError):
                    # If we can't find the original lead, just continue without CRM logging
                    pass
                
                return JsonResponse({'success': True, 'message': 'Status updated successfully'})
            else:
                return JsonResponse({'error': 'Invalid status'}, status=400)
                
        except LeadForward.DoesNotExist:
            return JsonResponse({'error': 'Forward record not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


def forward_activity_log(request, forward_id):
    """
    View activity log for a forwarded Google Sheets lead
    """
    try:
        current_employee = request.user.employee
        
        # Get forward record
        forward = LeadForward.objects.get(
            id=forward_id,
            forwarded_to=current_employee
        )
        
        # Get activity logs for this lead from CRM
        from apps.crm.models import LeadActivity
        # Create a dummy lead object for activity filtering
        dummy_lead = CrmLead.objects.get(id=int(forward.lead_id)) if forward.lead_id.isdigit() else None
        
        activities = LeadActivity.objects.filter(
            lead=dummy_lead
        ).select_related('created_by').order_by('-created_at')
        
        context = {
            'forward': forward,
            'activities': activities,
            'title': f'Activity Log - Lead {forward.lead_id}',
        }
        
        return render(request, 'accounts/forward_activity_log.html', context)
        
    except LeadForward.DoesNotExist:
        messages.error(request, 'Forward record not found.')
        return redirect('accounts:my_forwarded_leads')
    except Employee.DoesNotExist:
        messages.error(request, 'You need an employee profile to access this page.')
        return redirect('core:dashboard')


@login_required
def employee_organization_view(request):
    """
    View for employees to see their personalized team hierarchy
    Shows their manager, team leader, and team members
    """
    try:
        # Check if user is authenticated and has employee profile
        if not request.user.is_authenticated:
            messages.error(request, 'Please log in to access this page.')
            return redirect('core:dashboard')
        
        current_employee = request.user.employee
        
        # Get personalized team hierarchy
        manager = current_employee.reports_to if current_employee.reports_to and current_employee.reports_to.is_manager else None
        
        # Get team leader (for employees who report to a team leader)
        team_leader = None
        if current_employee.is_employee and current_employee.reports_to and current_employee.reports_to.is_team_leader:
            team_leader = current_employee.reports_to
        
        # Get team members (for team leaders and managers)
        team_members = []
        if current_employee.is_team_leader:
            team_members = Employee.objects.filter(
                reports_to=current_employee,
                employment_status='active'
            ).order_by('first_name', 'last_name')
        elif current_employee.is_manager:
            # Get team leaders under this manager
            team_leaders = Employee.objects.filter(
                reports_to=current_employee,
                role='team_leader',
                employment_status='active'
            )
            # Get all employees under those team leaders
            team_members = Employee.objects.filter(
                reports_to__in=team_leaders,
                employment_status='active'
            ).order_by('first_name', 'last_name')
        
        # Get all other employees for general view
        organization_employees = Employee.objects.filter(
            employment_status='active'
        ).order_by('first_name', 'last_name')
        
        # Apply role filter
        role_filter = request.GET.get('role', '')
        if role_filter and role_filter != 'all':
            organization_employees = organization_employees.filter(role=role_filter)
        
        # Apply search filter
        search_query = request.GET.get('search', '')
        if search_query:
            organization_employees = organization_employees.filter(
                models.Q(first_name__icontains=search_query) |
                models.Q(last_name__icontains=search_query) |
                models.Q(department__icontains=search_query) |
                models.Q(role__icontains=search_query) |
                models.Q(employee_id__icontains=search_query)
            )
        
        # Calculate base statistics (actual counts for each role)
        base_employees = Employee.objects.filter(
            employment_status='active'
        )
        
        # Calculate actual role counts (these should always show real totals)
        actual_total_count = base_employees.count()
        actual_manager_count = base_employees.filter(role='manager').count()
        actual_team_leader_count = base_employees.filter(role='team_leader').count()
        actual_employee_count = base_employees.filter(role='employee').count()
        
        # Calculate filtered statistics for current filter display
        if role_filter and role_filter != 'all':
            filtered_employees = base_employees.filter(role=role_filter)
        else:
            filtered_employees = base_employees
            
        if search_query:
            filtered_employees = filtered_employees.filter(
                models.Q(first_name__icontains=search_query) |
                models.Q(last_name__icontains=search_query) |
                models.Q(department__icontains=search_query) |
                models.Q(role__icontains=search_query) |
                models.Q(employee_id__icontains=search_query)
            )
        
        # Current filter results count
        filtered_count = filtered_employees.count()
        
        context = {
            'employees': organization_employees,
            'current_employee': current_employee,
            'manager': manager,
            'team_lead': team_leader,
            'team_members': team_members,
            'title': 'Organization Members',
            'total_count': actual_total_count,
            'manager_count': actual_manager_count,
            'team_leader_count': actual_team_leader_count,
            'employee_count': actual_employee_count,
            'filtered_count': filtered_count,
            'search_query': search_query,
            'role_filter': role_filter,
        }
        
        return render(request, 'accounts/employee_organization_view.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'You need an employee profile to access this page.')
        return redirect('core:dashboard')


@login_required
def employee_details_api(request, employee_id):
    """
    API endpoint to fetch employee details for modal display
    """
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get phone number if available
        phone = None
        if employee.phone_numbers.exists():
            phone = employee.phone_numbers.first().phone_number
        
        data = {
            'first_name': employee.first_name,
            'last_name': employee.last_name,
            'role': employee.role,
            'role_display': employee.get_role_display(),
            'department': employee.department,
            'department_display': employee.get_department_display(),
            'email': employee.user.email if employee.user else '',
            'phone': phone,
            'employment_status': employee.employment_status,
        }
        
        return JsonResponse(data)
        
    except Employee.DoesNotExist:
        return JsonResponse({'error': 'Employee not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def employee_list(request):

    """

    List employees with role-based filtering

    """

    try:

        current_employee = request.user.employee

        employees = current_employee.get_accessible_users()

    except Employee.DoesNotExist:

        # User doesn't have employee profile, redirect to dashboard or show message

        messages.error(request, 'You need an employee profile to access this page.')

        return redirect('core:dashboard')

    

    search_form = EmployeeSearchForm(request.GET)

    

    # Apply filters

    if search_form.is_valid():

        search = search_form.cleaned_data.get('search')

        department = search_form.cleaned_data.get('department')

        status = search_form.cleaned_data.get('status')

        employee_type = search_form.cleaned_data.get('employee_type')

        role = search_form.cleaned_data.get('role')

        

        if search:

            employees = employees.filter(

                Q(first_name__icontains=search) |

                Q(last_name__icontains=search) |

                Q(email__icontains=search) |

                Q(employee_id__icontains=search)

            )

        

        if department:

            employees = employees.filter(department=department)

        

        if status:

            employees = employees.filter(employment_status=status)

        

        if employee_type:

            employees = employees.filter(employee_type=employee_type)

        

        if role:

            employees = employees.filter(role=role)

    

    # Statistics

    total_employees = Employee.objects.count()

    active_employees = Employee.objects.filter(employment_status='active').count()

    onboarding_employees = Employee.objects.filter(employment_status='onboarding').count()

    managers_count = Employee.objects.filter(role='manager').count()

    

    context = {

        'employees': employees,

        'search_form': search_form,

        'total_employees': total_employees,

        'active_employees': active_employees,

        'onboarding_employees': onboarding_employees,

        'managers_count': managers_count,

    }

    return render(request, 'accounts/employee_list.html', context)





@staff_member_required

def employee_create(request):

    """

    Create a new employee

    """

    if request.method == 'POST':

        form = StaffForm(request.POST, user=request.user, is_create=True)

        if form.is_valid():

            try:

                employee = form.save()

                

                # Create default onboarding tasks

                default_tasks = [

                    "Complete employment paperwork",

                    "Set up email and system accounts",

                    "Review company policies",

                    "IT equipment setup",

                    "Office orientation tour",

                    "Introduction to team members",

                    "Set up workspace",

                ]

                

                for task_title in default_tasks:

                    OnboardingTask.objects.create(

                        employee=employee,

                        title=task_title,

                        status='pending'

                    )

                

                messages.success(request, f'Employee {employee.full_name} has been created successfully.')

                return redirect('accounts:staff_detail', pk=employee.pk)

            except Exception as e:

                messages.error(request, f'Error creating employee: {str(e)}')

        else:

            messages.error(request, 'Please correct the errors below.')

            print(f'Form errors: {form.errors}')  # Debug output

    else:

        form = StaffForm(user=request.user, is_create=True)

    

    return render(request, 'accounts/employee_form_professional.html', {'form': form, 'title': 'Create Employee'})





@can_view_user

def employee_detail(request, pk):

    """

    View employee details with permission check

    """

    employee = get_object_or_404(Employee, pk=pk)

    onboarding_tasks = employee.onboarding_tasks.all()

    documents = employee.documents.all()

    

    # Calculate onboarding progress

    total_tasks = onboarding_tasks.count()

    completed_tasks = onboarding_tasks.filter(status='completed').count()

    onboarding_progress = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

    

    # Get leads assigned to this employee (only for team leaders)
    leads = Lead.objects.none()
    lead_status_counts = {}
    date_filter = request.GET.get('date_filter')
    
    try:
        current_employee = request.user.employee
        
        # Only team leaders can see employee leads
        if current_employee.is_team_leader:
            leads = Lead.objects.filter(assigned_to=employee).order_by('-assigned_at')
            
            # Apply date filter if provided
            if date_filter:
                try:
                    from datetime import datetime
                    filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                    # Filter leads that were worked on this specific date
                    leads = leads.filter(
                        Q(assigned_at__date=filter_date) | 
                        Q(last_contacted__date=filter_date) |
                        Q(updated_at__date=filter_date)
                    ).distinct()
                except ValueError:
                    pass  # Invalid date format, ignore filter
            
            # Calculate lead status counts
            lead_status_counts = {
                'new': leads.filter(status='new').count(),
                'contacted': leads.filter(status='contacted').count(),
                'interested': leads.filter(status='interested').count(),
                'not_interested': leads.filter(status='not_interested').count(),
                'not_eligible': leads.filter(status='not_eligible').count(),
                'follow_up': leads.filter(status='follow_up').count(),
                'converted': leads.filter(status='converted').count(),
                'file_login': leads.filter(status='file_login').count(),
                'amount_disbursed': leads.filter(status='amount_disbursed').count(),
            }
    except Employee.DoesNotExist:
        pass

    

    # Check if current user can manage this employee

    try:

        current_employee = request.user.employee

        can_manage = current_employee.can_manage_user(employee)

    except Employee.DoesNotExist:

        can_manage = False

    

    # Pagination for leads
    from django.core.paginator import Paginator
    paginator = Paginator(leads, 20)
    page_number = request.GET.get('page')
    leads_page = paginator.get_page(page_number)

    

    context = {

        'employee': employee,

        'onboarding_tasks': onboarding_tasks,

        'documents': documents,

        'onboarding_progress': onboarding_progress,

        'can_manage': can_manage,

        'leads': leads_page,

        'lead_status_counts': lead_status_counts,

        'date_filter': date_filter,

    }

    return render(request, 'accounts/employee_profile.html', context)





@can_manage_user

def employee_update(request, pk):

    """

    Update employee information with permission check

    """

    employee = get_object_or_404(Employee, pk=pk)

    

    if request.method == 'POST':

        form = StaffForm(request.POST, instance=employee, user=request.user, is_create=False)

        if form.is_valid():

            form.save()

            messages.success(request, f'Employee {employee.full_name} has been updated successfully.')

            return redirect('accounts:staff_detail', pk=employee.pk)

    else:

        form = StaffForm(instance=employee, user=request.user, is_create=False)

    

    return render(request, 'accounts/employee_form_professional.html', {'form': form, 'title': 'Update Employee', 'employee': employee})





@staff_member_required

def employee_activate(request, pk):

    """

    Activate employee

    """

    employee = get_object_or_404(Employee, pk=pk)

    

    if request.method == 'POST':

        employee.employment_status = 'active'

        employee.user.is_active = True

        employee.user.save()

        employee.save()

        

        messages.success(request, f'Employee {employee.full_name} has been activated.')

        return redirect('accounts:staff_detail', pk=pk)

    

    return render(request, 'accounts/employee_confirm_activate.html', {'employee': employee})





@staff_member_required

def employee_deactivate(request, pk):

    """

    Deactivate employee

    """

    employee = get_object_or_404(Employee, pk=pk)

    

    if request.method == 'POST':

        employee.employment_status = 'inactive'

        employee.user.is_active = False

        employee.user.save()

        employee.save()

        

        messages.success(request, f'Employee {employee.full_name} has been deactivated.')

        return redirect('accounts:staff_detail', pk=pk)

    

    return render(request, 'accounts/employee_confirm_deactivate.html', {'employee': employee})





# Onboarding Task Views

@staff_member_required

def create_onboarding_task(request, employee_pk):

    """

    Create onboarding task for employee

    """

    employee = get_object_or_404(Employee, pk=employee_pk)

    

    if request.method == 'POST':

        form = OnboardingTaskForm(request.POST)

        if form.is_valid():

            task = form.save(commit=False)

            task.employee = employee

            task.save()

            messages.success(request, f'Onboarding task "{task.title}" has been created.')

            return redirect('accounts:staff_detail', pk=employee_pk)

    else:

        form = OnboardingTaskForm()

    

    return render(request, 'accounts/onboarding_task_form.html', {'form': form, 'employee': employee})





@staff_member_required

def update_onboarding_task(request, pk):

    """

    Update onboarding task status

    """

    task = get_object_or_404(OnboardingTask, pk=pk)

    

    if request.method == 'POST':

        form = OnboardingTaskForm(request.POST, instance=task)

        if form.is_valid():

            updated_task = form.save(commit=False)

            

            # Set completion details if marked as completed

            if updated_task.status == 'completed' and task.status != 'completed':

                updated_task.completed_date = timezone.now()

                updated_task.completed_by = request.user

            

            updated_task.save()

            messages.success(request, f'Task "{task.title}" has been updated.')

            return redirect('accounts:staff_detail', pk=task.employee.pk)

    

    return redirect('accounts:staff_detail', pk=task.employee.pk)





# Document Management Views

@staff_member_required

def upload_document(request, employee_pk):

    """

    Upload document for employee

    """

    employee = get_object_or_404(Employee, pk=employee_pk)

    

    if request.method == 'POST':

        form = DocumentUploadForm(request.POST, request.FILES)

        if form.is_valid():

            document = form.save(commit=False)

            document.employee = employee

            document.uploaded_by = request.user

            document.save()

            messages.success(request, f'Document "{document.title}" has been uploaded.')

            return redirect('accounts:staff_detail', pk=employee_pk)

    else:

        form = DocumentUploadForm()

    

    return render(request, 'accounts/document_upload.html', {'form': form, 'employee': employee})










# Dashboard Views

@staff_member_required

def hr_dashboard(request):

    """

    HR Dashboard with employee statistics

    """

    # Statistics

    total_employees = Employee.objects.count()

    active_employees = Employee.objects.filter(employment_status='active').count()

    onboarding_employees = Employee.objects.filter(employment_status='onboarding').count()

    managers_count = Employee.objects.filter(role__in=['manager', 'hr_manager', 'admin']).count()

    

    # Department breakdown

    dept_stats = {}

    for dept_code, dept_name in Employee.DEPARTMENTS:

        dept_stats[dept_name] = Employee.objects.filter(department=dept_code).count()

    

    # Role breakdown

    role_stats = {}

    for role_code, role_name in Employee.ROLES:

        role_stats[role_name] = Employee.objects.filter(role=role_code).count()

    

    # Recent hires

    recent_hires = Employee.objects.order_by('-created_at')[:5]

    

    # Pending onboarding tasks

    pending_tasks = OnboardingTask.objects.filter(status='pending').order_by('due_date')[:10]

    

    # Pending manager requests

    pending_requests = ManagerRequest.objects.filter(status='pending').order_by('-created_at')[:5]

    

    context = {

        'total_employees': total_employees,

        'active_employees': active_employees,

        'onboarding_employees': onboarding_employees,

        'managers_count': managers_count,

        'dept_stats': dept_stats,

        'role_stats': role_stats,

        'recent_hires': recent_hires,

        'pending_tasks': pending_tasks,

        'pending_requests': pending_requests,

    }

    return render(request, 'accounts/hr_dashboard.html', context)





# Manager Management Views

@login_required

def manager_list(request):

    """

    List all managers with role-based access

    """

    try:

        current_employee = request.user.employee

        

        if current_employee.is_superadmin:

            # SuperAdmin can see all managers

            managers = Employee.objects.filter(role='manager')

        elif current_employee.is_manager:

            # Manager can only see themselves

            managers = Employee.objects.filter(id=current_employee.id)

        else:

            # Other roles can't see managers

            managers = Employee.objects.none()

        

        context = {

            'managers': managers,

            'total_managers': managers.count(),

        }

        return render(request, 'accounts/manager_list.html', context)

    except Employee.DoesNotExist:

        messages.error(request, 'You need an employee profile to access this page.')

        return redirect('core:dashboard')





@staff_member_required

def create_manager_request(request, employee_pk):

    """

    Create a manager promotion request for an employee

    """

    employee = get_object_or_404(Employee, pk=employee_pk)

    

    if request.method == 'POST':

        form = ManagerRequestForm(request.POST, employee=employee)

        if form.is_valid():

            manager_request = form.save(commit=False)

            manager_request.employee = employee

            manager_request.save()

            

            messages.success(request, f'Manager promotion request for {employee.full_name} has been submitted.')

            return redirect('accounts:staff_detail', pk=employee_pk)

    else:

        form = ManagerRequestForm(employee=employee)

    

    return render(request, 'accounts/manager_request_form.html', {

        'form': form, 

        'employee': employee

    })





@staff_member_required

def manager_requests(request):

    """

    List all manager requests

    """

    requests = ManagerRequest.objects.all().order_by('-created_at')

    

    context = {

        'requests': requests,

        'pending_count': requests.filter(status='pending').count(),

    }

    return render(request, 'accounts/manager_requests.html', context)





@staff_member_required

def approve_manager_request(request, pk):

    """

    Approve a manager request

    """

    manager_request = get_object_or_404(ManagerRequest, pk=pk)

    

    if request.method == 'POST':

        # Update employee role

        employee = manager_request.employee

        employee.role = manager_request.requested_role

        

        # Update employee role - permissions are handled by the model's save method

        employee.save()

        

        # Update request status

        manager_request.status = 'approved'

        manager_request.approved_by = request.user

        manager_request.approved_at = timezone.now()

        manager_request.save()

        

        messages.success(request, f'{employee.full_name} has been promoted to {manager_request.get_requested_role_display()}.')

        return redirect('accounts:manager_requests')

    

    return render(request, 'accounts/manager_request_approve.html', {

        'manager_request': manager_request

    })





@staff_member_required

def reject_manager_request(request, pk):

    """

    Reject a manager request

    """

    manager_request = get_object_or_404(ManagerRequest, pk=pk)

    

    if request.method == 'POST':

        rejection_reason = request.POST.get('rejection_reason', '')

        

        manager_request.status = 'rejected'

        manager_request.rejection_reason = rejection_reason

        manager_request.save()

        

        messages.info(request, f'Manager request for {manager_request.employee.full_name} has been rejected.')

        return redirect('accounts:manager_requests')

    

    return render(request, 'accounts/manager_request_reject.html', {

        'manager_request': manager_request

    })





# Team Management Views

@role_required(['superadmin', 'manager'])

def team_list(request):

    """

    Show complete organizational hierarchy: Managers -> Team Leaders -> Employees

    """

    try:

        current_employee = request.user.employee

        

        if current_employee.is_superadmin:

            # SuperAdmin can see all managers and their teams

            managers = Employee.objects.filter(role='manager', employment_status='active')

        else:

            # Manager can only see their own team

            managers = Employee.objects.filter(id=current_employee.id, role='manager', employment_status='active')

            # If no managers found (edge case), ensure current manager is included
            if not managers.exists() and current_employee.role == 'manager':
                managers = Employee.objects.filter(id=current_employee.id)

        

        # Build hierarchy data

        hierarchy_data = []

        for manager in managers:

            team_leaders = manager.get_team_leaders()

            

            manager_data = {

                'manager': manager,

                'team_leaders': [],

                'total_team_members': 0

            }

            

            for team_leader in team_leaders:

                team_members = team_leader.get_team_members()

                

                team_leader_data = {

                    'team_leader': team_leader,

                    'team_members': team_members,

                    'member_count': team_members.count()

                }

                

                manager_data['team_leaders'].append(team_leader_data)

                manager_data['total_team_members'] += team_members.count()

            

            hierarchy_data.append(manager_data)

        

        context = {

            'hierarchy_data': hierarchy_data,

            'total_managers': len(hierarchy_data),

            'total_team_leaders': sum(len(m['team_leaders']) for m in hierarchy_data),

            'total_employees': sum(m['total_team_members'] for m in hierarchy_data),

        }

        

        return render(request, 'accounts/team_list.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'You need an employee profile to access this page.')

        return redirect('core:dashboard')





@role_required(['manager', 'hr_manager', 'admin'])

def create_team(request):

    """

    Create a new team

    """

    if request.method == 'POST':

        form = TeamForm(request.POST)

        if form.is_valid():

            team = form.save()

            messages.success(request, f'Team "{team.name}" has been created successfully.')

            return redirect('accounts:team_list')

    else:

        form = TeamForm()

    

    return render(request, 'accounts/team_form.html', {

        'form': form, 

        'title': 'Create Team'

    })





@role_required(['manager', 'hr_manager', 'admin'])

def update_team(request, pk):

    """

    Update an existing team

    """

    team = get_object_or_404(Team, pk=pk)

    

    if request.method == 'POST':

        form = TeamForm(request.POST, instance=team)

        if form.is_valid():

            form.save()

            messages.success(request, f'Team "{team.name}" has been updated successfully.')

            return redirect('accounts:team_list')

    else:

        form = TeamForm(instance=team)

    

    return render(request, 'accounts/team_form.html', {

        'form': form, 

        'title': f'Update Team: {team.name}',

        'team': team

    })





@role_required(['manager', 'hr_manager', 'admin'])

def team_detail(request, pk):

    """

    View team details and members

    """

    team = get_object_or_404(Team, pk=pk)

    members = team.get_members()

    

    context = {

        'team': team,

        'members': members,

        'member_count': members.count(),

    }

    return render(request, 'accounts/team_detail.html', context)





@login_required

def assign_employee_to_team(request, employee_pk):

    """

    Assign an employee to a team

    """

    employee = get_object_or_404(Employee, pk=employee_pk)

    

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to assign employees to teams.')

            return redirect('accounts:staff_detail', pk=employee_pk)

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('accounts:staff_detail', pk=employee_pk)

    

    if request.method == 'POST':

        team_leader_id = request.POST.get('team_leader')

        if team_leader_id:

            team_leader = get_object_or_404(Employee, pk=team_leader_id, role='team_leader')

            employee.team_leader = team_leader

            employee.save()

            messages.success(request, f'{employee.full_name} has been assigned to {team_leader.full_name}\'s team.')

        else:

            employee.team_leader = None

            employee.save()

            messages.info(request, f'{employee.full_name} has been removed from their team.')

        

        return redirect('accounts:staff_detail', pk=employee_pk)

    

    # Get available team leaders based on current user's role

    if current_employee.is_superadmin:

        team_leaders = Employee.objects.filter(role='team_leader', employment_status='active')

    elif current_employee.is_admin:

        team_leaders = Employee.objects.filter(role='team_leader', employment_status='active')

    elif current_employee.is_manager:

        # Manager can only assign to team leaders that report to them

        team_leaders = Employee.objects.filter(

            role='team_leader', 

            employment_status='active',

            reports_to=current_employee

        )

    else:

        team_leaders = Employee.objects.none()

    

    return render(request, 'accounts/assign_team.html', {

        'employee': employee,

        'team_leaders': team_leaders,

    })





@login_required

def employee_assignment(request):

    """

    Comprehensive employee assignment dashboard for managers and admins

    Shows all managers, their team leaders, and allows employee assignment

    """

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to access employee assignment.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    # Get managers based on current user's role

    if current_employee.is_superadmin:

        managers = Employee.objects.filter(role='manager', employment_status='active')

    elif current_employee.is_admin:

        managers = Employee.objects.filter(role='manager', employment_status='active')

    elif current_employee.is_manager:

        # Manager can only see themselves

        managers = Employee.objects.filter(id=current_employee.id)

    else:

        managers = Employee.objects.none()

    

    # Build hierarchy data

    hierarchy_data = []

    for manager in managers:

        # Get team leaders under this manager

        team_leaders = Employee.objects.filter(

            role='team_leader',

            employment_status='active',

            reports_to=manager

        )

        

        manager_data = {

            'manager': manager,

            'team_leaders': [],

            'total_team_members': 0,

            'unassigned_team_members': 0

        }

        

        for team_leader in team_leaders:

            # Get team members for this team leader

            team_members = Employee.objects.filter(

                team_leader=team_leader,

                employment_status='active'

            )

            

            # Get unassigned employees that could be assigned to this team leader

            # Only show employees who are not assigned to ANY team leader

            unassigned_employees = Employee.objects.filter(

                employment_status='active',

                role='employee',

                team_leader__isnull=True

            )

            

            team_leader_data = {

                'team_leader': team_leader,

                'team_members': team_members,

                'team_member_count': team_members.count(),

                'unassigned_employees': unassigned_employees,

                'unassigned_count': unassigned_employees.count()

            }

            

            manager_data['team_leaders'].append(team_leader_data)

            manager_data['total_team_members'] += team_members.count()

        

        # Get unassigned employees that could be assigned to any team leader under this manager

        manager_unassigned = Employee.objects.filter(

            employment_status='active',

            role='employee',

            team_leader__isnull=True

        )

        manager_data['unassigned_team_members'] = manager_unassigned.count()

        

        hierarchy_data.append(manager_data)

    

    # Get all unassigned employees for the quick assign section

    all_unassigned_employees = Employee.objects.filter(

        employment_status='active',

        role='employee',

        team_leader__isnull=True

    )

    

    # Get all team leaders for quick assignment

    all_team_leaders = Employee.objects.filter(

        role='team_leader',

        employment_status='active'

    )

    

    context = {

        'hierarchy_data': hierarchy_data,

        'all_unassigned_employees': all_unassigned_employees,

        'all_team_leaders': all_team_leaders,

        'total_managers': managers.count(),

        'total_unassigned': all_unassigned_employees.count(),

        'current_employee': current_employee,

    }

    

    return render(request, 'accounts/employee_assignment.html', context)





@login_required

@require_POST

def bulk_assign_employees(request):

    """

    Bulk assign employees to team leaders

    """

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                return JsonResponse({'success': False, 'message': 'Permission denied'})

            messages.error(request, 'Permission denied')

            return redirect('accounts:employee_assignment')

    except Employee.DoesNotExist:

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': 'Employee profile not found'})

        messages.error(request, 'Employee profile not found')

        return redirect('accounts:employee_assignment')

    

    employee_id = request.POST.get('employee_id')

    team_leader_id = request.POST.get('team_leader_id')

    

    if not employee_id or not team_leader_id:

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': 'Missing required parameters'})

        messages.error(request, 'Missing required parameters')

        return redirect('accounts:employee_assignment')

    

    try:

        employee = Employee.objects.get(pk=employee_id)

        team_leader = Employee.objects.get(pk=team_leader_id, role='team_leader')

        

        # Check if employee is already assigned to a team leader

        if employee.team_leader and employee.team_leader != team_leader:

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                return JsonResponse({'success': False, 'message': f'Employee is already assigned to {employee.team_leader.full_name}. Please unassign first.'})

            messages.error(request, f'Employee is already assigned to {employee.team_leader.full_name}. Please unassign first.')

            return redirect('accounts:employee_assignment')

        

        # Check permissions for this assignment

        if current_employee.is_manager:

            # Manager can only assign to team leaders under them

            if team_leader.reports_to != current_employee:

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                    return JsonResponse({'success': False, 'message': 'You can only assign to team leaders under your supervision'})

                messages.error(request, 'You can only assign to team leaders under your supervision')

                return redirect('accounts:employee_assignment')

        

        employee.team_leader = team_leader

        employee.save()

        

        success_message = f'{employee.full_name} has been assigned to {team_leader.full_name}'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': True, 'message': success_message})

        messages.success(request, success_message)

        return redirect('accounts:employee_assignment')

        

    except Employee.DoesNotExist:

        error_message = 'Employee or team leader not found'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': error_message})

        messages.error(request, error_message)

        return redirect('accounts:employee_assignment')

    except Exception as e:

        error_message = f'An error occurred: {str(e)}'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': error_message})

        messages.error(request, error_message)

        return redirect('accounts:employee_assignment')





@login_required

@require_POST

def quick_assign_to_team_leader(request):

    """

    Quick assign an employee to a specific team leader

    """

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                return JsonResponse({'success': False, 'message': 'Permission denied'})

            messages.error(request, 'Permission denied')

            return redirect('accounts:employee_assignment')

    except Employee.DoesNotExist:

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': 'Employee profile not found'})

        messages.error(request, 'Employee profile not found')

        return redirect('accounts:employee_assignment')

    

    employee_id = request.POST.get('employee_id')

    team_leader_id = request.POST.get('team_leader_id')

    

    if not employee_id or not team_leader_id:

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': 'Missing required parameters'})

        messages.error(request, 'Missing required parameters')

        return redirect('accounts:employee_assignment')

    

    try:

        employee = Employee.objects.get(pk=employee_id)

        team_leader = Employee.objects.get(pk=team_leader_id, role='team_leader')

        

        # Check if employee is already assigned to a team leader

        if employee.team_leader and employee.team_leader != team_leader:

            error_message = f'Employee is already assigned to {employee.team_leader.full_name}. Please unassign first.'

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                return JsonResponse({'success': False, 'message': error_message})

            messages.error(request, error_message)

            return redirect('accounts:employee_assignment')

        

        # Check permissions for this assignment

        if current_employee.is_manager:

            # Manager can only assign to team leaders under them

            if team_leader.reports_to != current_employee:

                error_message = 'You can only assign to team leaders under your supervision'

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                    return JsonResponse({'success': False, 'message': error_message})

                messages.error(request, error_message)

                return redirect('accounts:employee_assignment')

        

        employee.team_leader = team_leader

        employee.save()

        

        success_message = f'{employee.full_name} assigned to {team_leader.full_name}'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': True, 'message': success_message})

        messages.success(request, success_message)

        return redirect('accounts:employee_assignment')

        

    except Employee.DoesNotExist:

        error_message = 'Employee or team leader not found'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': error_message})

        messages.error(request, error_message)

        return redirect('accounts:employee_assignment')

    except Exception as e:

        error_message = f'An error occurred: {str(e)}'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return JsonResponse({'success': False, 'message': error_message})

        messages.error(request, error_message)

        return redirect('accounts:employee_assignment')





@login_required

@require_POST

def remove_from_team(request, employee_pk):

    """

    Remove an employee from their team leader

    """

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to remove employees from teams.')

            return redirect('accounts:employee_assignment')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('accounts:employee_assignment')

    

    try:

        employee = Employee.objects.get(pk=employee_pk)

        

        # Check permissions

        if current_employee.is_manager and employee.team_leader:

            # Manager can only remove from team leaders under them

            if employee.team_leader.reports_to != current_employee:

                messages.error(request, 'You can only remove employees from team leaders under your supervision.')

                return redirect('accounts:employee_assignment')

        

        team_leader_name = employee.team_leader.full_name if employee.team_leader else 'team'

        employee.team_leader = None

        employee.save()

        

        messages.success(request, f'{employee.full_name} has been removed from {team_leader_name}.')

        

    except Employee.DoesNotExist:

        messages.error(request, 'Employee not found.')

    

    return redirect('accounts:employee_assignment')





@login_required

def team_leader_assignment(request, team_leader_pk):

    """

    Show team leader details and allow admin/manager to assign employees to them

    """

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to assign employees to team leaders.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    team_leader = get_object_or_404(Employee, pk=team_leader_pk, role='team_leader')

    

    # Get all employees that can be assigned to this team leader

    # Only show employees who are not assigned to ANY team leader

    if current_employee.is_superadmin:

        # SuperAdmin can assign any employee

        assignable_employees = Employee.objects.filter(

            employment_status='active',

            role='employee',

            team_leader__isnull=True

        )

    elif current_employee.is_admin:

        # Admin can assign employees under their management

        assignable_employees = Employee.objects.filter(

            employment_status='active',

            role='employee',

            team_leader__isnull=True

        )

    elif current_employee.is_manager:

        # Manager can only assign employees that report to team leaders under them

        team_leaders_under_manager = Employee.objects.filter(

            role='team_leader',

            employment_status='active',

            reports_to=current_employee

        )

        assignable_employees = Employee.objects.filter(

            employment_status='active',

            role='employee',

            team_leader__isnull=True

        )

    else:

        assignable_employees = Employee.objects.none()

    

    # Get currently assigned employees

    assigned_employees = Employee.objects.filter(

        team_leader=team_leader,

        employment_status='active'

    )

    

    if request.method == 'POST':

        employee_id = request.POST.get('employee_id')

        action = request.POST.get('action')  # 'assign' or 'remove'

        

        if employee_id and action:

            employee = get_object_or_404(Employee, pk=employee_id)

            

            if action == 'assign':

                # Check if employee is already assigned to a team leader

                if employee.team_leader and employee.team_leader != team_leader:

                    messages.error(request, f'{employee.full_name} is already assigned to {employee.team_leader.full_name}. Please unassign first.')

                    return redirect('accounts:team_leader_assignment', team_leader_pk=team_leader_pk)

                elif employee.team_leader == team_leader:

                    messages.warning(request, f'{employee.full_name} is already assigned to this team leader.')

                    return redirect('accounts:team_leader_assignment', team_leader_pk=team_leader_pk)

                

                employee.team_leader = team_leader

                employee.save()

                messages.success(request, f'{employee.full_name} has been assigned to {team_leader.full_name}.')

            elif action == 'remove':

                employee.team_leader = None

                employee.save()

                messages.success(request, f'{employee.full_name} has been removed from {team_leader.full_name}\'s team.')

            

            return redirect('accounts:team_leader_assignment', team_leader_pk=team_leader_pk)

    

    context = {

        'team_leader': team_leader,

        'assignable_employees': assignable_employees,

        'assigned_employees': assigned_employees,

        'assignable_count': assignable_employees.count(),

        'assigned_count': assigned_employees.count(),

    }

    

    return render(request, 'accounts/team_leader_assignment.html', context)





@login_required

def my_team(request):

    """

    Show current user's team (for team leaders and managers)

    """

    try:

        employee = request.user.employee

        

        if employee.role == 'team_leader':

            team_members = employee.get_team_members()

            return render(request, 'accounts/my_team.html', {

                'team_leader': employee,

                'team_members': team_members,

                'member_count': team_members.count(),

            })

        elif employee.role == 'manager':

            # Manager can see all team leaders under them and their team members

            team_leaders_under_manager = Employee.objects.filter(

                role='team_leader',

                employment_status='active',

                reports_to=employee

            )

            

            # Get all team members under these team leaders

            all_team_members = Employee.objects.filter(

                team_leader__in=team_leaders_under_manager,

                employment_status='active'

            )

            

            # Group team members by team leader for easier display

            team_members_by_leader = {}

            for team_leader in team_leaders_under_manager:

                team_members_by_leader[team_leader] = all_team_members.filter(

                    team_leader=team_leader

                )

            

            return render(request, 'accounts/my_team.html', {

                'manager': employee,

                'team_leaders': team_leaders_under_manager,

                'team_members_by_leader': team_members_by_leader,

                'team_leader_count': team_leaders_under_manager.count(),

                'member_count': all_team_members.count(),

            })

        else:

            # Show employee's own team leader info

            if employee.team_leader:

                return render(request, 'accounts/my_team.html', {

                    'employee': employee,

                    'team_leader': employee.team_leader,

                })

            else:

                messages.info(request, 'You are not assigned to any team.')

                return redirect('core:dashboard')

                

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





# Excel Batch Management Views

import pandas as pd

import os

from django.http import JsonResponse

from django.core.paginator import Paginator

from django.utils import timezone

from .forms import ExcelUploadForm, ExcelBatchProcessForm

from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side





@login_required

def export_login_statistics_excel(request):

    """Export login statistics to Excel file with optional filters"""

    # Check permissions

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin or request.user.employee.is_manager):

        messages.error(request, 'You do not have permission to export login statistics.')

        return redirect('accounts:login_statistics_dashboard')

    

    # Get all statistics with related data

    statistics = LoginStatistics.objects.select_related('manager', 'team_leader', 'updated_by')

    

    # Apply filters from GET parameters

    applied_filters = []

    

    # Date filters

    date_from = request.GET.get('date_from')

    date_to = request.GET.get('date_to')

    

    if date_from:

        statistics = statistics.filter(date__gte=date_from)

        applied_filters.append(f"Date From: {date_from}")

    

    if date_to:

        statistics = statistics.filter(date__lte=date_to)

        applied_filters.append(f"Date To: {date_to}")

    

    # Manager filter

    manager_id = request.GET.get('manager')

    if manager_id:

        try:

            manager = Employee.objects.get(id=manager_id)

            statistics = statistics.filter(manager=manager)

            applied_filters.append(f"Manager: {manager.full_name}")

        except Employee.DoesNotExist:

            pass

    

    # Team leader filter

    team_leader_id = request.GET.get('team_leader')

    if team_leader_id:

        try:

            team_leader = Employee.objects.get(id=team_leader_id)

            statistics = statistics.filter(team_leader=team_leader)

            applied_filters.append(f"Team Leader: {team_leader.full_name}")

        except Employee.DoesNotExist:

            pass

    

    # Order by date (newest first)

    statistics = statistics.order_by('-date', 'manager', 'team_leader')

    

    # Create workbook and worksheet

    wb = Workbook()

    ws = wb.active

    ws.title = "Login Statistics"

    

    # Define styles

    header_font = Font(bold=True, color="FFFFFF")

    header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")

    border = Border(

        left=Side(border_style="thin", color="000000"),

        right=Side(border_style="thin", color="000000"),

        top=Side(border_style="thin", color="000000"),

        bottom=Side(border_style="thin", color="000000")

    )

    

    # Add filters info at the top if any filters were applied

    row_num = 1

    if applied_filters:

        ws.cell(row=row_num, column=1, value="Applied Filters:")

        for filter_info in applied_filters:

            row_num += 1

            ws.cell(row=row_num, column=1, value=f"• {filter_info}")

        row_num += 2  # Add space before headers

    

    # Set column headers

    headers = [

        'Date', 'Manager', 'Team Leader', 'File Count', 'Login Count', 

        'Total Count', 'Notes', 'Updated By', 'Updated At'

    ]

    

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=row_num, column=col_num, value=header)

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = Alignment(horizontal="center")

        cell.border = border

    

    # Set column widths

    column_widths = [12, 25, 25, 12, 12, 12, 30, 20, 20]

    for i, width in enumerate(column_widths, 1):

        ws.column_dimensions[chr(64+i)].width = width

    

    # Add data rows

    data_start_row = row_num + 1

    for stat in statistics:

        row_data = [

            stat.date.strftime('%Y-%m-%d'),

            stat.manager.full_name if stat.manager else '',

            stat.team_leader.full_name if stat.team_leader else '',

            stat.file_count,

            stat.login_count,

            stat.total_count,

            stat.notes or '',

            stat.updated_by.username if stat.updated_by else '',

            stat.updated_at.strftime('%Y-%m-%d %H:%M') if stat.updated_at else ''

        ]

        

        for col_num, value in enumerate(row_data, 1):

            cell = ws.cell(row=data_start_row, column=col_num, value=value)

            cell.border = border

            # Center align numeric columns

            if col_num in [4, 5, 6]:

                cell.alignment = Alignment(horizontal="center")

        

        data_start_row += 1

    

    # Create filename with filters info

    filename_parts = ["login_statistics"]

    if applied_filters:

        filename_parts.append("filtered")

    filename_parts.append(timezone.now().strftime('%Y%m%d_%H%M%S'))

    filename = "_".join(filename_parts) + ".xlsx"

    

    # Create response

    response = HttpResponse(

        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    

    # Save workbook to response

    wb.save(response)

    return response





@login_required

def excel_upload(request):

    """Upload Excel file and create a new batch"""

    # Check permissions - only admin and superadmin can upload

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to upload Excel files.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    if request.method == 'POST':

        form = ExcelUploadForm(request.POST, request.FILES)

        print(f"Form is valid: {form.is_valid()}")

        if form.errors:

            print(f"Form errors: {form.errors}")

        if form.is_valid():

            # Create batch

            batch = form.save(commit=False)

            batch.created_by = request.user

            batch.save()

            

            # Process Excel file

            try:

                excel_file = batch.excel_file

                file_path = excel_file.path

                

                print(f"Processing file: {file_path}")

                print(f"File exists: {os.path.exists(file_path)}")

                

                # Read Excel file

                if excel_file.name.endswith('.csv'):

                    df = pd.read_csv(file_path)

                else:

                    df = pd.read_excel(file_path)

                

                print(f"DataFrame shape: {df.shape}")

                print(f"Columns: {list(df.columns)}")

                

                # Create leads from Excel data

                leads = []

                for index, row in df.iterrows():

                    # Convert row to dict with proper datetime handling
                    row_data = {}
                    for col in df.columns:
                        value = row[col]
                        # Handle datetime objects in keys/values
                        if isinstance(value, (datetime, pd.Timestamp)):
                            row_data[str(col)] = value.isoformat() if hasattr(value, 'isoformat') else str(value)
                        elif pd.isna(value):
                            row_data[str(col)] = None
                        else:
                            row_data[str(col)] = value

                    

                    # Skip empty rows (check if all values are NaN or empty)

                    if all(str(val).strip() in ['', 'nan', 'NaN', 'None'] for val in row_data.values()):

                        continue

                    

                    # Create lead with all data stored in JSON

                    lead = Lead(source=batch.batch_name)

                    lead.set_data(row_data)

                    

                    # Create batch row for reference

                    batch_row = ExcelBatchRow(

                        batch=batch,

                        row_number=index + 2,

                    )

                    batch_row.set_row_data(row_data)

                    batch_row.status = 'processed'

                    batch_row.employee_created = current_employee

                    batch_row.processed_at = timezone.now()

                    batch_row.save()

                    

                    # Link lead to batch row and set creator

                    lead.batch_row = batch_row

                    lead.updated_by = current_employee

                    lead.save()

                    

                    leads.append(lead)

                

                # Update batch status

                batch.total_rows = len(leads)

                batch.processed_rows = len(leads)

                batch.status = 'completed'

                batch.completed_at = timezone.now()

                batch.save()

                

                messages.success(request, f'Excel file uploaded successfully! {batch.total_rows} leads created from batch "{batch.batch_name}"')

                return redirect('accounts:excel_batch_detail', pk=batch.pk)

                

            except Exception as e:

                print(f"Error processing Excel file: {str(e)}")

                import traceback

                traceback.print_exc()

                batch.status = 'failed'

                batch.error_message = str(e)

                batch.save()

                messages.error(request, f'Error processing Excel file: {str(e)}')

                return redirect('accounts:excel_upload')

        else:

            messages.error(request, 'Please correct the errors below.')

    else:

        form = ExcelUploadForm()

    

    return render(request, 'accounts/excel_upload.html', {'form': form})





@login_required

def excel_batch_list(request):

    """List all Excel batches"""

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to view Excel batches.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    batches = ExcelBatch.objects.all().order_by('-created_at')

    

    # Filter by status

    status_filter = request.GET.get('status')

    if status_filter:

        batches = batches.filter(status=status_filter)

    

    # Pagination

    paginator = Paginator(batches, 10)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    context = {

        'page_obj': page_obj,

        'status_filter': status_filter,

        'total_batches': batches.count(),

    }

    

    return render(request, 'accounts/excel_batch_list.html', context)





@login_required

def excel_batch_detail(request, pk):

    """View and manage a specific Excel batch"""

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to view Excel batches.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    batch = get_object_or_404(ExcelBatch, pk=pk)

    batch_rows = batch.batch_rows.all().order_by('row_number')

    

    # Filter rows by status

    status_filter = request.GET.get('row_status')

    if status_filter:

        batch_rows = batch_rows.filter(status=status_filter)

    

    # Pagination for rows

    paginator = Paginator(batch_rows, 20)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    # Status counts

    status_counts = {

        'pending': batch.batch_rows.filter(status='pending').count(),

        'processed': batch.batch_rows.filter(status='processed').count(),

        'failed': batch.batch_rows.filter(status='failed').count(),

        'skipped': batch.batch_rows.filter(status='skipped').count(),

    }

    

    # Get sample data for preview

    sample_rows = batch.batch_rows.all()[:5]

    

    if request.method == 'POST':

        form = ExcelBatchProcessForm(batch, request.POST)

        if form.is_valid():

            action = form.cleaned_data['action']

            

            if action == 'cancel':

                batch.status = 'cancelled'

                batch.save()

                messages.success(request, f'Batch "{batch.batch_name}" has been cancelled.')

                return redirect('accounts:excel_batch_list')

            

            elif action in ['process', 'retry_failed', 'skip_errors']:

                # Process the batch

                try:

                    process_excel_batch(batch, action)

                    messages.success(request, f'Batch processing started for "{batch.batch_name}"')

                except Exception as e:

                    batch.status = 'failed'

                    batch.error_message = str(e)

                    batch.save()

                    messages.error(request, f'Error processing batch: {str(e)}')

                

                return redirect('accounts:excel_batch_detail', pk=batch.pk)

    else:

        form = ExcelBatchProcessForm(batch)

    

    context = {

        'batch': batch,

        'page_obj': page_obj,

        'form': form,

        'status_counts': status_counts,

        'status_filter': status_filter,

        'sample_rows': sample_rows,

    }

    

    return render(request, 'accounts/excel_batch_detail.html', context)





@login_required

def process_excel_batch(batch, action):

    """Process Excel batch and create employees"""

    from django.contrib.auth.models import User

    import uuid

    

    batch.status = 'processing'

    batch.save()

    

    # Get rows to process

    if action == 'retry_failed':

        rows_to_process = batch.batch_rows.filter(status='failed')

    elif action == 'skip_errors':

        rows_to_process = batch.batch_rows.filter(status__in=['pending', 'failed'])

    else:  # process all

        rows_to_process = batch.batch_rows.filter(status='pending')

    

    processed_count = 0

    failed_count = 0

    

    for batch_row in rows_to_process:

        try:

            row_data = batch_row.get_row_data()  # Use new method to get JSON data

            

            # Map Excel columns to employee fields

            # No required fields - process whatever data is available

            employee_data = {

                'first_name': str(row_data.get('first_name', '')).strip() or 'Unknown',

                'last_name': str(row_data.get('last_name', '')).strip() or 'Unknown',

                'email': str(row_data.get('email', '')).strip() or f'user_{batch_row.row_number}@example.com',

                'phone': str(row_data.get('phone', '')).strip(),

                'role': str(row_data.get('role', 'employee')).strip().lower() or 'employee',

                'position': str(row_data.get('position', '')).strip() or 'Employee',

                'department': str(row_data.get('department', '')).strip() or 'General',

                'employment_status': 'active'

            }

            

            # Generate unique username and email if needed

            base_email = employee_data['email']

            if employee_data['email'] == f'user_{batch_row.row_number}@example.com' or User.objects.filter(email=employee_data['email']).exists():

                # Generate unique email

                employee_data['email'] = f'user_{batch_row.row_number}_{batch.pk}@example.com'

            

            # Create unique username

            username = f"user_{batch_row.row_number}_{batch.pk}"

            base_username = username

            counter = 1

            while User.objects.filter(username=username).exists():

                username = f"{base_username}_{counter}"

                counter += 1

            

            # Create user

            user = User.objects.create_user(

                username=username,

                email=employee_data['email'],

                password=str(uuid.uuid4())[:8],  # Random password

                first_name=employee_data['first_name'],

                last_name=employee_data['last_name']

            )

            

            # Create employee

            employee = Employee.objects.create(

                user=user,

                **employee_data

            )

            

            batch_row.status = 'processed'

            batch_row.employee_created = employee

            batch_row.processed_at = timezone.now()

            batch_row.save()

            processed_count += 1

            

        except Exception as e:

            batch_row.status = 'failed'

            batch_row.error_message = str(e)

            batch_row.save()

            failed_count += 1

    

    # Update batch status

    batch.processed_rows += processed_count

    batch.failed_rows += failed_count

    

    if batch.processed_rows + batch.failed_rows >= batch.total_rows:

        batch.status = 'completed'

        batch.completed_at = timezone.now()

    

    batch.save()





@login_required

def leads_by_status(request, status):

    """

    View to display leads filtered by status

    """

    # Validate status parameter

    valid_statuses = [choice[0] for choice in Lead.LEAD_STATUS_CHOICES]

    if status not in valid_statuses:

        messages.error(request, 'Invalid status filter.')

        return redirect('accounts:lead_list')

    

    try:

        # Get leads filtered by status

        leads = Lead.objects.filter(status=status).select_related('assigned_to', 'batch_row__batch')

        

        # Apply role-based filtering

        if not request.user.employee.is_superadmin and not request.user.employee.is_admin:

            leads = leads.filter(assigned_to=request.user.employee)

            

    except Employee.DoesNotExist:

        # User doesn't have employee profile

        leads = Lead.objects.none()

    

    # Order by most recently updated

        leads = leads.order_by('-updated_at')

    

    # Pagination

    paginator = Paginator(leads, 20)

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    # Calculate status counts for context

    status_counts = {}

    for stat in valid_statuses:

        status_counts[stat] = Lead.objects.filter(status=stat).count()

    

    context = {

        'page_obj': page_obj,

        'status_filter': status,

        'status_counts': status_counts,

        'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

    }

    

    return render(request, 'accounts/leads_by_status.html', context)





@login_required

def lead_list(request):

    """View all leads"""

    try:

        current_employee = request.user.employee

        

        # Filter leads based on user role

        if current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager:

            leads = Lead.objects.all()

        else:

            # Employees can only see their assigned leads

            leads = Lead.objects.filter(assigned_to=current_employee)

            

            # Additional security: redirect if employee tries to access all leads page

            if not leads.exists() and request.path == '/accounts/leads/':

                messages.info(request, 'You have been redirected to your assigned leads.')

                return redirect('accounts:my_leads')

        

        # Calculate statistics on ALL leads (before filtering)

        status_counts = {}

        for status_value, status_label in Lead.LEAD_STATUS_CHOICES:

            status_counts[status_value] = leads.filter(status=status_value).count()

        

        # Filter by status

        status_filter = request.GET.get('status', '')

        if status_filter:

            leads = leads.filter(status=status_filter)

        

        # Search functionality

        search_query = request.GET.get('search', '')

        if search_query:

            # Search in JSON data fields

            import json

            matching_leads = []

            for lead in leads:

                data = lead.get_data()

                if any(search_query.lower() in str(value).lower() for value in data.values() if value is not None):

                    matching_leads.append(lead.id)

            leads = leads.filter(id__in=matching_leads)

        

        # Pagination

        from django.core.paginator import Paginator

        paginator = Paginator(leads, 20)

        page_number = request.GET.get('page')

        page_obj = paginator.get_page(page_number)

        

        context = {

            'page_obj': page_obj,

            'status_filter': status_filter,

            'search_query': search_query,

            'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

            'status_counts': status_counts,

        }

        

        return render(request, 'accounts/lead_list.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required

def lead_detail(request, pk):

    """View lead details"""

    lead = get_object_or_404(Lead, pk=pk)

    

    try:

        current_employee = request.user.employee

        

        # Check permissions
        if current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager:
            # Admins and managers can view any lead
            pass
        elif current_employee.is_team_leader:
            # Team leaders can view leads assigned to their team members
            if lead.assigned_to:
                # Check if the lead is assigned to the team leader or one of their team members
                team_members = Employee.objects.filter(team_leader=current_employee, employment_status='active')
                if lead.assigned_to and lead.assigned_to not in team_members and lead.assigned_to != current_employee:
                    messages.error(request, 'You do not have permission to view this lead.')
                    employee_id = request.GET.get('employee')
                    if employee_id:
                        return redirect('accounts:employee_detail', pk=employee_id)
                    return redirect('accounts:lead_list')
            else:
                # Lead not assigned to anyone, team leaders can't view
                messages.error(request, 'You do not have permission to view this lead.')
                employee_id = request.GET.get('employee')
                if employee_id:
                    return redirect('accounts:employee_detail', pk=employee_id)
                return redirect('accounts:lead_list')
        elif lead.assigned_to == current_employee:
            # Employee can view their own leads
            pass
        else:
            messages.error(request, 'You do not have permission to view this lead.')
            return redirect('accounts:lead_list')

        

        if request.method == 'POST':
            # Update lead status
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            if new_status in [choice[0] for choice in Lead.LEAD_STATUS_CHOICES]:
                old_status = lead.status
                
                # Log the status change before updating
                from .models import log_lead_activity
                log_lead_activity(
                    lead=lead,
                    user=request.user,
                    action_type='status_changed',
                    old_value=old_status,
                    new_value=new_status,
                    description=f"Changed lead status from '{old_status}' to '{new_status}' for {lead.get_full_name()}",
                    request=request
                )
                
                lead.status = new_status
                lead.notes = notes
                lead.last_contacted = timezone.now()
                lead.save()
                
                messages.success(request, f'Lead status updated to {new_status}.')
                
                # Find next available NEW lead for this employee (only NEW status leads)
                next_lead = Lead.objects.filter(
                    assigned_to=current_employee,
                    status='new'  # ONLY get NEW status leads
                ).exclude(pk=lead.pk).order_by('created_at').first()
                
                if next_lead:
                    messages.success(request, f'Lead updated! Moving to next NEW lead: {next_lead.get_full_name()}')
                    return redirect('accounts:lead_detail', pk=next_lead.pk)
                else:
                    messages.info(request, 'No more NEW leads available! All leads have been processed.')
                    return redirect('accounts:my_leads')

        context = {

            'lead': lead,

            'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

            'additional_data': lead.get_data(),

        }

        

        return render(request, 'accounts/lead_detail.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required

@login_required
def lead_activity_logs(request):
    """View and filter lead activity logs for admin"""
    try:
        current_employee = request.user.employee
        
        # Only admins can view activity logs
        if not (current_employee.is_superadmin or current_employee.is_admin):
            messages.error(request, 'You do not have permission to view activity logs.')
            return redirect('core:dashboard')
        
        # Get filter parameters
        lead_id_filter = request.GET.get('lead_id', '')
        action_type_filter = request.GET.get('action_type', '')
        user_filter = request.GET.get('user', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        # Start with all activity logs
        activities = LeadActivityLog.objects.all()
        
        # Apply filters
        if lead_id_filter:
            activities = activities.filter(lead_id=lead_id_filter)
        
        if action_type_filter:
            activities = activities.filter(action_type=action_type_filter)
        
        if user_filter:
            activities = activities.filter(user_id=user_filter)
        
        if start_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                activities = activities.filter(created_at__date__gte=start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                from datetime import datetime
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                activities = activities.filter(created_at__date__lte=end_date_obj)
            except ValueError:
                pass
        
        # Order by latest first
        activities = activities.select_related('lead', 'user').order_by('-created_at')
        
        # Get filter options for dropdowns
        action_types = LeadActivityLog.ACTION_TYPES
        users = User.objects.filter(is_active=True).order_by('username')
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(activities, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'activities': page_obj,
            'action_types': action_types,
            'users': users,
            'lead_id_filter': lead_id_filter,
            'action_type_filter': action_type_filter,
            'user_filter': user_filter,
            'start_date': start_date,
            'end_date': end_date,
        }
        
        return render(request, 'accounts/lead_activity_logs.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('core:dashboard')


@login_required
def lead_activity_detail(request, activity_id):
    """View details of a specific lead activity"""
    try:
        current_employee = request.user.employee
        
        # Only admins can view activity details
        if not (current_employee.is_superadmin or current_employee.is_admin):
            messages.error(request, 'You do not have permission to view activity details.')
            return redirect('core:dashboard')
        
        activity = get_object_or_404(LeadActivityLog, pk=activity_id)
        
        context = {
            'activity': activity,
        }
        
        return render(request, 'accounts/lead_activity_detail.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('core:dashboard')


def assign_leads(request):
    """Assign leads to employees with bulk assignment option"""
    try:
        current_employee = request.user.employee
        
        # Only admins and managers can assign leads
        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):
            messages.error(request, 'You do not have permission to assign leads.')
            return redirect('accounts:lead_list')
        
        if request.method == 'POST':
            # Check if this is bulk assignment
            employee_ids = request.POST.getlist('employee_ids')
            lead_ids = request.POST.getlist('lead_ids')
            
            if employee_ids and lead_ids:
                # Bulk assignment mode - assign 20 leads to each selected employee
                leads_per_employee = 20
                assigned_count = 0
                
                for employee_id in employee_ids:
                    assigned_employee = get_object_or_404(Employee, pk=employee_id)
                    
                    # Get unassigned leads for this employee (avoid duplicates across employees)
                    available_leads = Lead.objects.filter(assigned_to__isnull=True)[:leads_per_employee]
                    
                    for lead in available_leads:
                        lead.assigned_to = assigned_employee
                        lead.assigned_by = current_employee
                        lead.assigned_at = timezone.now()
                        lead.save()
                        
                        # Log the assignment
                        from .models import log_lead_activity
                        log_lead_activity(
                            lead=lead,
                            user=current_employee,
                            action_type='assigned',
                            old_value=None,
                            new_value=assigned_employee.get_full_name(),
                            description=f"Assigned lead '{lead.get_full_name()}' to {assigned_employee.get_full_name()}",
                            request=request
                        )
                        assigned_count += 1
                
                messages.success(request, f'Successfully assigned {assigned_count} leads to {len(employee_ids)} employees (20 each).')
                return redirect('accounts:lead_list')
            
            # Single employee assignment (existing logic)
            employee_id = request.POST.get('employee_id')
            lead_ids = request.POST.getlist('lead_ids')
            
            if lead_ids and employee_id:
                assigned_employee = get_object_or_404(Employee, pk=employee_id)
                leads = Lead.objects.filter(pk__in=lead_ids)
                
                for lead in leads:
                    lead.assigned_to = assigned_employee
                    lead.assigned_by = current_employee
                    lead.assigned_at = timezone.now()
                    lead.save()
                    
                    # Log the assignment
                    from .models import log_lead_activity
                    log_lead_activity(
                        lead=lead,
                        user=current_employee,
                        action_type='assigned',
                        old_value=None,
                        new_value=assigned_employee.get_full_name(),
                        description=f"Assigned lead '{lead.get_full_name()}' to {assigned_employee.get_full_name()}",
                        request=request
                    )
                
                messages.success(request, f'{len(leads)} leads assigned to {assigned_employee.get_full_name()}.')
                return redirect('accounts:lead_list')

        # Get unassigned leads for selection with date range and datasource filtering
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        datasource_filter = request.GET.get('datasource', '')
        
        unassigned_leads = Lead.objects.filter(assigned_to__isnull=True)
        
        # Apply date range filter
        if start_date:
            try:
                from datetime import datetime
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                unassigned_leads = unassigned_leads.filter(created_at__date__gte=start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                from datetime import datetime
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                unassigned_leads = unassigned_leads.filter(created_at__date__lte=end_date_obj)
            except ValueError:
                pass
        
        # Apply datasource filter
        if datasource_filter:
            unassigned_leads = unassigned_leads.filter(source=datasource_filter)
        
        # Order by creation date (newest first)
        unassigned_leads = unassigned_leads.order_by('-created_at')
        
        # Get unique datasources for filter dropdown
        datasources = Lead.objects.filter(
            assigned_to__isnull=True,
            source__isnull=False
        ).exclude(source='').values_list('source', flat=True).distinct().order_by('source')
        
        employees = Employee.objects.filter(
            role__in=['employee', 'team_leader'],
            employment_status='active'
        )
        
        
        context = {
            'unassigned_leads': unassigned_leads,
            'employees': employees,
            'datasources': datasources,
            'selected_datasource': datasource_filter,
            'start_date': start_date,
            'end_date': end_date,
        }

        

        return render(request, 'accounts/assign_leads.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')


@login_required
def emi_calculator(request):
    """EMI Calculator for team leaders and employees"""
    try:
        current_employee = request.user.employee
        
        # Only team leaders and employees can access EMI calculator
        if current_employee.role not in ['employee', 'team_leader']:
            messages.error(request, 'You do not have permission to access the EMI calculator.')
            return redirect('core:dashboard')
        
        calculation = None
        form_data = {}
        
        if request.method == 'POST':
            # Get form data
            principal = request.POST.get('principal')
            interest_rate = request.POST.get('interest_rate')
            tenure_months = request.POST.get('tenure_months')
            
            form_data = {
                'principal': principal,
                'interest_rate': interest_rate,
                'tenure_months': tenure_months,
            }
            
            # Validate inputs
            try:
                principal = float(principal) if principal else 0
                interest_rate = float(interest_rate) if interest_rate else 0
                tenure_months = int(tenure_months) if tenure_months else 0
                
                if principal <= 0 or interest_rate <= 0 or tenure_months <= 0:
                    messages.error(request, 'Please enter valid values greater than zero.')
                else:
                    # EMI calculation formula: EMI = [P x R x (1+R)^N] / [(1+R)^N-1]
                    # Where:
                    # P = Principal amount
                    # R = Monthly interest rate (annual rate / 12 / 100)
                    # N = Number of monthly installments
                    
                    monthly_rate = interest_rate / 12 / 100
                    emi = (principal * monthly_rate * pow(1 + monthly_rate, tenure_months)) / (pow(1 + monthly_rate, tenure_months) - 1)
                    
                    total_amount = emi * tenure_months
                    total_interest = total_amount - principal
                    
                    calculation = {
                        'principal': principal,
                        'interest_rate': interest_rate,
                        'tenure_months': tenure_months,
                        'monthly_emi': round(emi, 2),
                        'total_amount': round(total_amount, 2),
                        'total_interest': round(total_interest, 2),
                        'monthly_rate_percent': round(monthly_rate * 100, 4),
                    }
                    
                    messages.success(request, f'EMI calculation completed! Monthly EMI: ₹{calculation["monthly_emi"]}')
                    
            except (ValueError, TypeError):
                messages.error(request, 'Please enter valid numeric values.')
        
        context = {
            'calculation': calculation,
            'form_data': form_data,
        }
        
        return render(request, 'accounts/emi_calculator.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Employee profile not found.')
        return redirect('core:dashboard')


@login_required
def my_leads(request):

    """View leads assigned to current user"""

    try:

        current_employee = request.user.employee

        leads = Lead.objects.filter(assigned_to=current_employee)


        # Calculate statistics on ALL leads (before filtering)

        status_counts = {}

        for status_value, status_label in Lead.LEAD_STATUS_CHOICES:

            status_counts[status_value] = leads.filter(status=status_value).count()


        # Filter by status

        status_filter = request.GET.get('status', '')

        if status_filter:

            leads = leads.filter(status=status_filter)
        
        # Filter by datasource
        datasource_filter = request.GET.get('datasource', '')
        
        if datasource_filter:
            leads = leads.filter(source=datasource_filter)
        
        # Filter by single date
        assigned_date = request.GET.get('assigned_date', '')
        
        if assigned_date:
            try:
                assigned_date_obj = datetime.strptime(assigned_date, '%Y-%m-%d').date()
                leads = leads.filter(assigned_at__date=assigned_date_obj)
            except ValueError:
                pass
        
        # Get unique datasources for filter dropdown
        datasources = Lead.objects.filter(
            assigned_to=current_employee,
            source__isnull=False
        ).exclude(source='').values_list('source', flat=True).distinct().order_by('source')
        
        # Pagination

        from django.core.paginator import Paginator

        paginator = Paginator(leads, 20)

        page_number = request.GET.get('page')

        page_obj = paginator.get_page(page_number)

        

        context = {

            'page_obj': page_obj,

            'status_filter': status_filter,

            'datasource_filter': datasource_filter,

            'assigned_date': assigned_date,

            'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

            'status_counts': status_counts,

            'datasources': datasources,

        }

        

        return render(request, 'accounts/my_leads.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required

def api_leads_for_deletion(request):

    """API endpoint to get leads for deletion"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):

        return JsonResponse({'error': 'Permission denied'}, status=403)

    

    leads = Lead.objects.all()

    

    # Apply filters

    status_filter = request.GET.get('status')

    assignment_filter = request.GET.get('assignment')

    batch_filter = request.GET.get('batch')

    

    if status_filter:

        leads = leads.filter(status=status_filter)

    

    if assignment_filter == 'assigned':

        leads = leads.filter(assigned_to__isnull=False)

    elif assignment_filter == 'unassigned':

        leads = leads.filter(assigned_to__isnull=True)

    

    if batch_filter and batch_filter != '':

        leads = leads.filter(batch_row__batch__name=batch_filter)

    

    # Prepare response data

    leads_data = []

    for lead in leads:

        leads_data.append({

            'id': lead.id,

            'full_name': lead.get_full_name(),

            'email': lead.get_email(),

            'phone': lead.get_phone(),

            'company': lead.get_company(),

            'status': lead.status,

            'data': lead.get_data(),  # Include all uploaded data

            'assigned_to': {

                'first_name': lead.assigned_to.first_name if lead.assigned_to else None,

                'last_name': lead.assigned_to.last_name if lead.assigned_to else None,

            } if lead.assigned_to else None,

            'batch_name': lead.batch_row.batch.batch_name if lead.batch_row and lead.batch_row.batch else 'Manual'

        })

    

    return JsonResponse({'leads': leads_data})





@login_required

def api_employee_leads(request, employee_id):

    """API endpoint to get leads for a specific employee"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_manager or request.user.employee.is_hr):

        return JsonResponse({'error': 'Permission denied'}, status=403)

    

    try:

        employee = get_object_or_404(Employee, pk=employee_id)

        leads = Lead.objects.filter(assigned_to=employee)

        

        leads_data = []

        for lead in leads:

            lead_data = lead.get_data()

            leads_data.append({

                'id': lead.id,

                'first_name': lead_data.get('first_name', ''),

                'last_name': lead_data.get('last_name', ''),

                'email': lead_data.get('email', ''),

                'phone': lead_data.get('phone', ''),

                'company': lead_data.get('company', ''),

                'position': lead_data.get('position', ''),

                'status': lead.status,

                'assigned_at': lead.assigned_at.isoformat() if lead.assigned_at else None,

            })

        

        return JsonResponse({'leads': leads_data})

        

    except Employee.DoesNotExist:

        return JsonResponse({'error': 'Employee not found'}, status=404)





@login_required

def unassign_leads(request):

    """Unassign leads from employees"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_manager or request.user.employee.is_hr):

        messages.error(request, 'You do not have permission to unassign leads.')

        return redirect('accounts:lead_list')

    

    if request.method == 'POST':

        employee_id = request.POST.get('employee_id')

        lead_ids = request.POST.getlist('lead_ids')

        

        if employee_id and lead_ids:

            employee = get_object_or_404(Employee, pk=employee_id)

            leads = Lead.objects.filter(pk__in=lead_ids, assigned_to=employee)

            

            count = leads.count()

            leads.update(assigned_to=None, assigned_at=None, assigned_by=None)

            

            messages.success(request, f'Successfully unassigned {count} lead(s) from {employee.get_full_name}.')

        else:

            messages.error(request, 'Please select an employee and at least one lead.')

        

        return redirect('accounts:lead_list')

    

    # GET request - show unassign interface

    datasource_filter = request.GET.get('datasource', '')
    current_employee = request.user.employee

    if datasource_filter:
        if current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager:
            # Admins and managers can see all employees with leads from any source
            employees = Employee.objects.filter(
                role__in=['employee', 'team_leader'], 
                employment_status='active'
            ).filter(assigned_leads__assigned_to__isnull=False, assigned_leads__source=datasource_filter).distinct()
        else:
            # Team leaders can only see themselves
            employees = Employee.objects.filter(
                id=current_employee.id,
                employment_status='active'
            ).filter(assigned_leads__assigned_to__isnull=False, assigned_leads__source=datasource_filter).distinct()
    else:
        if current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager:
            # Admins and managers can see all employees
            employees = Employee.objects.filter(
                role__in=['employee', 'team_leader'], 
                employment_status='active'
            )
        else:
            # Team leaders can only see themselves
            employees = Employee.objects.filter(
                id=current_employee.id,
                employment_status='active'
            )

    # Get unique datasources for filter dropdown - based on user role
    if current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager:
        datasources = Lead.objects.filter(
            assigned_to__isnull=False,
            source__isnull=False
        ).exclude(source='').values_list('source', flat=True).distinct().order_by('source')
    else:
        # Team leaders can only see their own datasources
        datasources = Lead.objects.filter(
            assigned_to=current_employee,
            assigned_to__isnull=False,
            source__isnull=False
        ).exclude(source='').values_list('source', flat=True).distinct().order_by('source')

    context = {
        'employees': employees,
        'datasources': datasources,
        'selected_datasource': datasource_filter,
    }

    return render(request, 'accounts/unassign_leads.html', context)





@login_required

def delete_leads(request):

    """Delete leads permanently"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_manager or request.user.employee.is_hr):

        messages.error(request, 'You do not have permission to delete leads.')

        return redirect('accounts:lead_list')

    

    if request.method == 'POST':

        if 'delete' in request.POST:

            lead_ids = request.POST.getlist('lead_ids')

            

            if lead_ids:

                leads = Lead.objects.filter(pk__in=lead_ids)

                count = leads.count()

                

                # Store info for message

                batch_info = []

                for lead in leads:

                    if lead.batch_row and lead.batch_row.batch:

                        batch_info.append(f"Batch '{lead.batch_row.batch.batch_name}'")

                

                leads.delete()

                

                if batch_info:

                    messages.success(request, f'Successfully deleted {count} lead(s) from {len(set(batch_info))} batch(es).')

                else:

                    messages.success(request, f'Successfully deleted {count} lead(s).')

                

                return redirect('accounts:lead_list')

            else:

                messages.error(request, 'Please select at least one lead to delete.')

        

        elif 'filter' in request.POST:

            # Handle filtering

            status_filter = request.POST.get('status_filter')

            assignment_filter = request.POST.get('assignment_filter')

            

            leads = Lead.objects.all()

            

            if status_filter:

                leads = leads.filter(status=status_filter)

            

            if assignment_filter == 'assigned':

                leads = leads.filter(assigned_to__isnull=False)

            elif assignment_filter == 'unassigned':

                leads = leads.filter(assigned_to__isnull=True)

            

            context = {

                'filtered_leads': leads,

                'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

                'status_filter': status_filter,

                'assignment_filter': assignment_filter,

            }

            return render(request, 'accounts/delete_leads.html', context)

    

    # GET request - show all leads initially

    leads = Lead.objects.all()

    

    context = {

        'filtered_leads': leads,

        'lead_status_choices': Lead.LEAD_STATUS_CHOICES,

    }

    return render(request, 'accounts/delete_leads.html', context)





@login_required

def delete_excel_batch(request, pk):

    """Delete an Excel batch and all associated leads"""

    # Check permissions

    try:

        current_employee = request.user.employee

        if not (current_employee.is_superadmin or current_employee.is_admin or current_employee.is_manager):

            messages.error(request, 'You do not have permission to delete Excel batches.')

            return redirect('core:dashboard')

    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    

    batch = get_object_or_404(ExcelBatch, pk=pk)

    

    if request.method == 'POST':

        batch_name = batch.batch_name

        

        # Count leads that will be deleted

        leads_to_delete = Lead.objects.filter(batch_row__batch=batch)

        leads_count = leads_to_delete.count()

        

        # Delete leads first (they will cascade to batch rows)

        leads_to_delete.delete()

        

        # Delete the batch (this will also delete remaining batch rows)

        batch.delete()

        

        messages.success(request, f'Batch "{batch_name}" and {leads_count} associated leads have been deleted successfully.')

        return redirect('accounts:excel_batch_list')

    

    # Show confirmation with lead count

    leads_count = Lead.objects.filter(batch_row__batch=batch).count()

    return render(request, 'accounts/excel_batch_delete.html', {'batch': batch, 'leads_count': leads_count})





@login_required

def user_activity_logs(request):

    """View for displaying user activity logs with Indian Standard Time"""

    # Check permissions - only admin/superadmin/manager can access

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin or request.user.employee.is_manager):

        messages.error(request, 'You do not have permission to access user activity logs.')

        return redirect('core:dashboard')

    

    # Get filter parameters

    user_filter = request.GET.get('user', '')

    date_filter = request.GET.get('date', '')

    status_filter = request.GET.get('status', '')

    

    # Start with all logs

    logs = UserActivityLog.objects.all()

    

    # Apply filters

    if user_filter:

        logs = logs.filter(user__username__icontains=user_filter)

    

    if date_filter:

        try:

            filter_date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()

            logs = logs.filter(login_time__date=filter_date)

        except ValueError:

            pass

    

    if status_filter == 'active':

        logs = logs.filter(is_active_session=True)

    elif status_filter == 'completed':

        logs = logs.filter(is_active_session=False)

    

    # Order by most recent first

    logs = logs.select_related('user').order_by('-login_time')

    

    # Pagination

    from django.core.paginator import Paginator

    paginator = Paginator(logs, 50)  # 50 logs per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    context = {

        'page_obj': page_obj,

        'user_filter': user_filter,

        'date_filter': date_filter,

        'status_filter': status_filter,

        'total_logs': logs.count(),

        'active_sessions_count': UserActivityLog.objects.filter(is_active_session=True).count(),

    }

    

    return render(request, 'accounts/user_activity_logs.html', context)





@login_required

def employee_performance_dashboard(request):

    """Dashboard for tracking employee performance and lead interactions"""

    # Check permissions - only admin/superadmin/manager can access

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin or request.user.employee.is_manager):

        messages.error(request, 'You do not have permission to access performance dashboard.')

        return redirect('core:dashboard')

    

    # Get filter parameters

    employee_filter = request.GET.get('employee', '')

    date_from = request.GET.get('date_from', '')

    date_to = request.GET.get('date_to', '')

    action_filter = request.GET.get('action', '')

    

    # Start with all activity logs

    activities = LeadActivityLog.objects.all()

    

    # Apply filters

    if employee_filter:

        activities = activities.filter(user__username__icontains=employee_filter)

    

    if date_from:

        try:

            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()

            activities = activities.filter(created_at__date__gte=from_date)

        except ValueError:

            pass

    

    if date_to:

        try:

            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()

            activities = activities.filter(created_at__date__lte=to_date)

        except ValueError:

            pass

    

    if action_filter:

        activities = activities.filter(action_type=action_filter)

    

    # Order by most recent first

    activities = activities.select_related('user', 'lead').order_by('-created_at')

    

    # Calculate performance metrics

    from django.db.models import Count, Q

    

    # Get all users with activities for comprehensive tracking

    all_users_with_activities = User.objects.filter(

        lead_activities__isnull=False

    ).distinct().select_related('employee')

    

    # Employee performance summary with detailed tracking

    employee_stats = {}

    user_activity_details = {}

    

    for activity in activities:

        username = activity.user.username

        if username not in employee_stats:

            employee_stats[username] = {

                'user': activity.user,

                'total_actions': 0,

                'leads_created': 0,

                'leads_updated': 0,

                'leads_assigned': 0,

                'leads_unassigned': 0,

                'status_changes': 0,

                'leads_converted': 0,

                'leads_deleted': 0,

                'unique_leads': set(),

                'last_activity': None,

                'first_activity': None,

                'activities_by_type': {},

                'daily_activity': {},

                'recent_activities': []

            }

            user_activity_details[username] = []

        

        stats = employee_stats[username]

        stats['total_actions'] += 1

        stats['unique_leads'].add(activity.lead.pk)

        

        # Track first and last activity

        if not stats['first_activity'] or activity.created_at < stats['first_activity']:

            stats['first_activity'] = activity.created_at

        if not stats['last_activity'] or activity.created_at > stats['last_activity']:

            stats['last_activity'] = activity.created_at

        

        # Track activities by type

        action_type = activity.get_action_type_display()

        stats['activities_by_type'][action_type] = stats['activities_by_type'].get(action_type, 0) + 1

        

        # Track daily activity

        activity_date = activity.created_at.date().strftime('%Y-%m-%d')

        stats['daily_activity'][activity_date] = stats['daily_activity'].get(activity_date, 0) + 1

        

        # Add to detailed activity list

        user_activity_details[username].append({

            'id': activity.id,

            'action_type': activity.action_type,

            'action_display': action_type,

            'lead_name': activity.lead.get_full_name(),

            'lead_email': activity.lead.get_email(),

            'lead_company': activity.lead.get_company(),

            'description': activity.description,

            'old_value': activity.old_value,

            'new_value': activity.new_value,

            'timestamp': activity.created_at,

            'timestamp_ist': activity.get_created_at_ist(),

            'date': activity.created_at.date(),

            'time': activity.created_at.time()

        })

        

        # Keep only recent 10 activities for display

        if len(stats['recent_activities']) < 10:

            stats['recent_activities'].append({

                'action': activity.action_type,

                'lead': activity.lead.get_full_name(),

                'description': activity.description,

                'timestamp': activity.get_created_at_ist()

            })

        

        # Count specific action types

        if activity.action_type == 'created':

            stats['leads_created'] += 1

        elif activity.action_type == 'updated':

            stats['leads_updated'] += 1

        elif activity.action_type == 'assigned':

            stats['leads_assigned'] += 1

        elif activity.action_type == 'unassigned':

            stats['leads_unassigned'] += 1

        elif activity.action_type == 'status_changed':

            stats['status_changes'] += 1

        elif activity.action_type == 'converted':

            stats['leads_converted'] += 1

        elif activity.action_type == 'deleted':

            stats['leads_deleted'] += 1

    

    # Convert sets to counts and calculate additional metrics

    for username, stats in employee_stats.items():

        stats['unique_leads_count'] = len(stats['unique_leads'])

        del stats['unique_leads']

        

        # Calculate activity span (days between first and last activity)

        if stats['first_activity'] and stats['last_activity']:

            stats['activity_span_days'] = (stats['last_activity'].date() - stats['first_activity'].date()).days + 1

            stats['avg_activities_per_day'] = round(stats['total_actions'] / stats['activity_span_days'], 2)

        else:

            stats['activity_span_days'] = 0

            stats['avg_activities_per_day'] = 0

        

        # Sort recent activities by timestamp

        stats['recent_activities'].sort(key=lambda x: x['timestamp'], reverse=True)

        

        # Sort detailed activities by timestamp

        user_activity_details[username].sort(key=lambda x: x['timestamp'], reverse=True)

    

    # Get comprehensive user statistics including inactive users

    all_employees = Employee.objects.select_related('user').all()

    comprehensive_user_stats = {}

    

    for employee in all_employees:

        username = employee.user.username

        if username in employee_stats:

            comprehensive_user_stats[username] = employee_stats[username]

        else:

            # Add inactive users with zero stats

            comprehensive_user_stats[username] = {

                'user': employee.user,

                'total_actions': 0,

                'leads_created': 0,

                'leads_updated': 0,

                'leads_assigned': 0,

                'leads_unassigned': 0,

                'status_changes': 0,

                'leads_converted': 0,

                'leads_deleted': 0,

                'unique_leads_count': 0,

                'last_activity': None,

                'first_activity': None,

                'activities_by_type': {},

                'daily_activity': {},

                'recent_activities': [],

                'activity_span_days': 0,

                'avg_activities_per_day': 0,

                'is_active': False

            }

    

    # Sort employees by total actions

    sorted_employees = sorted(

        comprehensive_user_stats.items(),

        key=lambda x: x[1]['total_actions'],

        reverse=True

    )

    

    # Pagination

    from django.core.paginator import Paginator

    paginator = Paginator(activities, 50)  # 50 activities per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)

    

    # Calculate overall statistics

    total_activities = activities.count()

    active_employees = len([s for s in comprehensive_user_stats.values() if s['total_actions'] > 0])

    total_employees = len(comprehensive_user_stats)

    # Calculate phone assignment statistics

    from .models import EmployeePhone

    total_phone_assignments = EmployeePhone.objects.count()

    active_phone_assignments = EmployeePhone.objects.filter(is_active=True).count()

    inactive_phone_assignments = total_phone_assignments - active_phone_assignments

    employees_with_phones = EmployeePhone.objects.filter(is_active=True).values('employee').distinct().count()

    phone_assignment_stats = {

        'total_assignments': total_phone_assignments,

        'active_assignments': active_phone_assignments,

        'inactive_assignments': inactive_phone_assignments,

        'employees_with_phones': employees_with_phones,

    }

    

    # Get activity trends for the last 30 days

    from datetime import timedelta

    thirty_days_ago = timezone.now().date() - timedelta(days=30)

    daily_trends = {}

    

    for i in range(30):

        date = thirty_days_ago + timedelta(days=i)

        date_str = date.strftime('%Y-%m-%d')

        daily_trends[date_str] = activities.filter(created_at__date=date).count()

    

    context = {

        'page_obj': page_obj,

        'employee_filter': employee_filter,

        'date_from': date_from,

        'date_to': date_to,

        'action_filter': action_filter,

        'employee_stats': sorted_employees[:10],  # Top 10 performers

        'all_employee_stats': comprehensive_user_stats,  # All employees with detailed stats

        'user_activity_details': user_activity_details,  # Detailed activity logs for each user

        'total_activities': total_activities,

        'total_employees': total_employees,

        'active_employees': active_employees,

        'top_performer': sorted_employees[0] if sorted_employees else None,

        'daily_trends': daily_trends,

        'activity_types': dict(LeadActivityLog.ACTION_TYPES),

        'phone_assignment_stats': phone_assignment_stats,

    }

    

    return render(request, 'accounts/employee_performance_dashboard.html', context)





@login_required

def login_statistics_dashboard(request):

    """Dashboard for viewing and managing login statistics"""

    # Check permissions - only admin/superadmin/manager can access

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin or request.user.employee.is_manager):

        messages.error(request, 'You do not have permission to access login statistics.')

        return redirect('core:dashboard')

    

    # Handle form submission for new statistics

    if request.method == 'POST':

        if 'add_stats' in request.POST:

            form = LoginStatisticsForm(request.POST, user=request.user)

            if form.is_valid():

                stats = form.save(commit=False)

                stats.updated_by = request.user

                stats.save()

                messages.success(request, f'Login statistics added successfully for {stats.manager.full_name} on {stats.date}.')

                return redirect('accounts:login_statistics_dashboard')

            else:

                # Debug: Print form errors to console

                print(f"Form errors: {form.errors}")

                print(f"Form data: {request.POST}")

                messages.error(request, f'Please correct the errors: {form.errors}')

        else:

            form = LoginStatisticsForm(user=request.user)

    else:

        form = LoginStatisticsForm(user=request.user)

    # Handle filtering
    filter_form = LoginStatisticsFilterForm(request.GET or None, user=request.user)
    
    # Base queryset - filter for current manager's team
    statistics = LoginStatistics.objects.filter(
        Q(manager=request.user.employee) |
        Q(team_leader__reports_to=request.user.employee)
    )
    
    # Apply filters
    if filter_form and filter_form.is_valid():
        cleaned_data = filter_form.cleaned_data
        
        if cleaned_data.get('manager'):
            # Only show managers that are under current manager (if applicable)
            if request.user.employee.is_manager:
                statistics = statistics.filter(manager=cleaned_data['manager'])
            else:
                # Admin/superadmin can see any manager
                pass
                
        if cleaned_data.get('team_leader'):
            statistics = statistics.filter(team_leader=cleaned_data['team_leader'])
            
        if cleaned_data.get('date_from'):
            statistics = statistics.filter(date__gte=cleaned_data['date_from'])
            
        if cleaned_data.get('date_to'):
            statistics = statistics.filter(date__lte=cleaned_data['date_to'])

    # Optimize query for better performance

    statistics = statistics.select_related('manager', 'team_leader', 'updated_by').order_by('-date', 'manager', 'team_leader')

    

    # Calculate summary statistics

    total_files = statistics.aggregate(total=models.Sum('file_count'))['total'] or 0

    total_logins = statistics.aggregate(total=models.Sum('login_count'))['total'] or 0

    total_all = statistics.aggregate(total=models.Sum('total_count'))['total'] or 0

    

    # Group by manager for summary with optimized query

    manager_stats = {}

    for stat in statistics:

        manager_name = stat.manager.full_name

        if manager_name not in manager_stats:

            manager_stats[manager_name] = {

                'files': 0, 'logins': 0, 'total': 0, 'team_leaders': []

            }

        manager_stats[manager_name]['files'] += stat.file_count

        manager_stats[manager_name]['logins'] += stat.login_count

        manager_stats[manager_name]['total'] += stat.total_count

        if stat.team_leader and stat.team_leader.full_name not in manager_stats[manager_name]['team_leaders']:

            manager_stats[manager_name]['team_leaders'].append(stat.team_leader.full_name)

    

    # Get today's statistics for quick reference

    today = timezone.now().date()

    today_stats = statistics.filter(date=today)

    today_files = today_stats.aggregate(total=models.Sum('file_count'))['total'] or 0

    today_logins = today_stats.aggregate(total=models.Sum('login_count'))['total'] or 0

    

    context = {

        'form': form,

        'filter_form': filter_form,

        'statistics': statistics,

        'manager_stats': manager_stats,

        'summary': {

            'total_files': total_files,

            'total_logins': total_logins,

            'total_all': total_all,

            'total_entries': statistics.count()

        },

        'today_stats': {

            'today_files': today_files,

            'today_logins': today_logins,

            'today_entries': today_stats.count()

        }

    }

    return render(request, 'accounts/login_statistics_dashboard.html', context)





@login_required

def delete_login_statistics(request, pk):

    """Delete login statistics"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):

        messages.error(request, 'You do not have permission to delete login statistics.')

        return redirect('accounts:login_statistics_dashboard')

    # Rest of your code remains the same

    # ... rest of your code ...





@login_required

def start_work(request):

    """Start work for the current employee"""

    try:

        employee = request.user.employee



        # Check if user is admin/superadmin - they should only monitor, not participate

        if employee.is_superadmin or employee.is_admin:

            messages.warning(request, 'Administrators cannot start work. You can only monitor attendance.')

            return redirect('accounts:attendance_admin')



        success, message = employee.start_work()



        if success:

            messages.success(request, message)

        else:

            messages.error(request, message)



        return redirect('core:dashboard')



    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required

def end_work(request):

    """End work for the current employee"""

    try:

        employee = request.user.employee



        # Check if user is admin/superadmin - they should only monitor, not participate

        if employee.is_superadmin or employee.is_admin:

            messages.warning(request, 'Administrators cannot end work. You can only monitor attendance.')

            return redirect('accounts:attendance_admin')



        success, message = employee.end_work()



        if success:

            messages.success(request, message)

        else:

            messages.error(request, message)



        return redirect('core:dashboard')



    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required

def employee_attendance_stats(request, employee_id=None):

    """View detailed attendance statistics for an employee"""

    try:

        # Get employee

        if employee_id:

            employee = get_object_or_404(Employee, pk=employee_id)

        else:

            employee = request.user.employee



        # Check permissions - only admin/superadmin can view other employees' stats

        if employee_id and not (request.user.employee.is_superadmin or request.user.employee.is_admin):

            if employee != request.user.employee:

                messages.error(request, 'You do not have permission to view this attendance data.')

                return redirect('core:dashboard')



        # Get date range (last 30 days by default)

        date_from = request.GET.get('date_from', '')

        date_to = request.GET.get('date_to', '')



        if not date_from:

            from datetime import timedelta

            date_from = (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')



        # Get attendance records

        attendances = Attendance.objects.filter(employee=employee)



        if date_from:

            try:

                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()

                attendances = attendances.filter(date__gte=from_date)

            except ValueError:

                pass



        if date_to:

            try:

                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()

                attendances = attendances.filter(date__lte=to_date)

            except ValueError:

                pass



        attendances = attendances.order_by('-date')



        # Calculate statistics

        stats = employee.get_attendance_stats(

            start_date=from_date if date_from else None,

            end_date=to_date if date_to else None

        )



        # Get recent violations

        recent_violations = Violation.objects.filter(

            employee=employee

        ).order_by('-violation_time')[:10]



        # Get attendance timeline for the period

        attendance_timeline = []

        for attendance in attendances[:30]:  # Last 30 records

            attendance_timeline.append({

                'date': attendance.date,

                'status': attendance.get_status_display(),

                'start_time': attendance.start_time,

                'end_time': attendance.end_time,

                'total_hours': attendance.total_hours,

                'overtime_hours': attendance.overtime_hours,

                'breaks': attendance.breaks.count(),

            })



        context = {

            'employee': employee,

            'attendances': attendances,

            'stats': stats,

            'recent_violations': recent_violations,

            'attendance_timeline': attendance_timeline,

            'date_from': date_from,

            'date_to': date_to,

        }



        return render(request, 'accounts/employee_attendance_stats.html', context)



    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')



@login_required
def attendance_admin(request):
    """Admin view for managing all attendance records"""
    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr or request.user.employee.is_manager):
        messages.error(request, 'You do not have permission to access attendance admin.')
        return redirect('core:dashboard')


    # Get filter parameters

    employee_filter = request.GET.get('employee', '')

    status_filter = request.GET.get('status', '')

    date_from = request.GET.get('date_from', '')

    date_to = request.GET.get('date_to', '')



    # Get attendance records based on user role

    if request.user.employee.is_manager and not (request.user.employee.is_admin or request.user.employee.is_superadmin):

        # Managers see only their team members and team leaders

        attendances = Attendance.objects.filter(

            Q(employee__reports_to=request.user.employee) |

            Q(employee__team_leader=request.user.employee) |

            Q(employee=request.user.employee)

        ).select_related('employee', 'employee__user')

    else:

        # Admin/HR/Superadmin see all attendance records

        attendances = Attendance.objects.all().select_related('employee', 'employee__user')



    if employee_filter:

        attendances = attendances.filter(

            Q(employee__user__username__icontains=employee_filter) |

            Q(employee__first_name__icontains=employee_filter) |

            Q(employee__last_name__icontains=employee_filter)

        )



    if status_filter:

        attendances = attendances.filter(status=status_filter)



    if date_from:

        try:

            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()

            attendances = attendances.filter(date__gte=from_date)

        except ValueError:

            pass



    if date_to:

        try:

            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()

            attendances = attendances.filter(date__lte=to_date)

        except ValueError:

            pass



    # Order by most recent first

    attendances = attendances.order_by('-date')



    # Pagination

    from django.core.paginator import Paginator

    paginator = Paginator(attendances, 20)  # 20 records per page

    page_number = request.GET.get('page')

    page_obj = paginator.get_page(page_number)



    # Get statistics

    total_records = attendances.count()

    present_count = attendances.filter(status='present').count()

    late_count = attendances.filter(status='late').count()

    absent_count = attendances.filter(status='absent').count()



    # Get recent violations

    recent_violations = Violation.objects.filter(

        resolved=False

    ).select_related('employee', 'employee__user').order_by('-violation_time')[:10]



    context = {

        'page_obj': page_obj,

        'employee_filter': employee_filter,

        'status_filter': status_filter,

        'date_from': date_from,

        'date_to': date_to,

        'total_records': total_records,

        'present_count': present_count,

        'late_count': late_count,

        'absent_count': absent_count,

        'recent_violations': recent_violations,

    }



    return render(request, 'accounts/attendance_admin.html', context)





@login_required

def resolve_violation(request, violation_id):

    """Resolve a violation"""

    if not (request.user.employee.is_superadmin or request.user.employee.is_admin or request.user.employee.is_hr):

        messages.error(request, 'You do not have permission to resolve violations.')

        return redirect('accounts:attendance_admin')



    violation = get_object_or_404(Violation, pk=violation_id)



    if request.method == 'POST':

        review_notes = request.POST.get('review_notes', '')



        violation.resolved = True

        violation.resolved_by = request.user.employee

        violation.resolved_at = timezone.now()

        violation.review_notes = review_notes

        violation.save()



        messages.success(request, f'Violation resolved: {violation.get_violation_type_display()}')

        return redirect('accounts:attendance_admin')



    context = {

        'violation': violation,

    }



    return render(request, 'accounts/resolve_violation.html', context)





@login_required

def start_break(request, break_type='other'):

    """Start a break for the current employee"""

    try:

        employee = request.user.employee



        # Check if user is admin/superadmin/HR - they should only monitor, not participate
        if employee.is_superadmin or employee.is_admin or employee.is_hr:
            messages.warning(request, 'Administrators cannot take breaks. You can only monitor attendance.')
            return redirect('accounts:attendance_admin')



        success, message = employee.start_break(break_type)

        if success:
            # Get the active break to pass to dashboard
            active_break = employee.current_attendance.breaks.filter(end_time__isnull=True).first() if employee.current_attendance else None

            # Redirect to dashboard with break popup context
            from django.shortcuts import render
            from django.contrib import messages as django_messages

            # Add success message
            django_messages.success(request, message)

            # Render dashboard with break popup
            return render(request, 'core/dashboard.html', {
                'show_break_popup': True,
                'active_break': active_break,
                'break_started': True,
            })

        else:
            messages.error(request, message)

        return redirect('core:dashboard')



    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')





@login_required
def end_break(request):

    """End the current break for the employee"""

    try:

        employee = request.user.employee



        # Check if user is admin/superadmin/HR - they should only monitor, not participate
        if employee.is_superadmin or employee.is_admin or employee.is_hr:
            messages.warning(request, 'Administrators cannot take breaks. You can only monitor attendance.')
            return redirect('accounts:attendance_admin')



        success, message = employee.end_break()



        if success:

            messages.success(request, message)

        else:

            messages.error(request, message)



        return redirect('core:dashboard')



    except Employee.DoesNotExist:

        messages.error(request, 'Employee profile not found.')

        return redirect('core:dashboard')

    



@login_required

def manager_details(request, manager_name):

    """Detailed view for manager statistics"""

    # Check permissions - only admin/superadmin can access

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin or request.user.employee.is_manager):

        messages.error(request, 'You do not have permission to access manager details.')

        return redirect('accounts:login_statistics_dashboard')

    

    # Find the manager by name - split manager_name and search both first_name and last_name

    name_parts = manager_name.split()

    manager_query = Employee.objects.filter(role__in=['manager', 'admin', 'superadmin'])

    

    if len(name_parts) >= 2:

        manager_query = manager_query.filter(

            first_name__icontains=name_parts[0],

            last_name__icontains=name_parts[1]

        )

    else:

        manager_query = manager_query.filter(

            first_name__icontains=name_parts[0]

        )

    

    try:

        manager = manager_query.first()

        if not manager:

            messages.error(request, f'Manager "{manager_name}" not found.')

            return redirect('accounts:login_statistics_dashboard')

        

        # Get manager's team statistics

        team_leaders = Employee.objects.filter(reports_to=manager, role='team_leader', employment_status='active')

        team_members = Employee.objects.filter(

            reports_to__in=team_leaders, 

            role='employee', 

            employment_status='active'

        )

        

        # Calculate team performance metrics

        total_team_members = team_members.count()

        active_team_members = team_members.filter(employment_status='active').count()

        

        # Get recent activity for manager's team
        team_activity = LeadActivityLog.objects.filter(
            user__in=[member.user for member in team_members]
        ).select_related('user', 'lead').order_by('-created_at')[:20]

        

        context = {

            'manager': manager,

            'team_leaders': team_leaders,

            'team_members': team_members,

            'total_team_members': total_team_members,

            'active_team_members': active_team_members,

            'team_activity': team_activity,

            'team_leaders_count': team_leaders.count(),

        }

        

        return render(request, 'accounts/manager_details.html', context)

        

    except Employee.DoesNotExist:

        messages.error(request, 'Manager not found.')

        return redirect('accounts:login_statistics_dashboard')





@login_required

def update_login_statistics(request, pk):

    """Update existing login statistics"""

    # Check permissions

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin):

        messages.error(request, 'You do not have permission to update login statistics.')

        return redirect('accounts:login_statistics_dashboard')

    

    stats = get_object_or_404(LoginStatistics, pk=pk)

    

    if request.method == 'POST':

        form = LoginStatisticsForm(request.POST, instance=stats, user=request.user)

        if form.is_valid():

            updated_stats = form.save(commit=False)

            updated_stats.updated_by = request.user

            updated_stats.save()

            messages.success(request, 'Login statistics updated successfully.')

            return redirect('accounts:login_statistics_dashboard')

    else:

        form = LoginStatisticsForm(instance=stats, user=request.user)

    

    return render(request, 'accounts/update_login_statistics.html', {

        'form': form,

        'stats': stats

    })





@login_required

def delete_login_statistics(request, pk):

    """Delete login statistics"""

    # Check permissions

    if not (request.user.employee.is_admin or request.user.employee.is_superadmin):

        messages.error(request, 'You do not have permission to delete login statistics.')

        return redirect('accounts:login_statistics_dashboard')

    

    stats = get_object_or_404(LoginStatistics, pk=pk)

    

    if request.method == 'POST':

        stats.delete()

        messages.success(request, 'Login statistics deleted successfully.')

        return redirect('accounts:login_statistics_dashboard')

    

    return render(request, 'accounts/delete_login_statistics.html', {'stats': stats})





@login_required

def my_leaves(request):

    """View for employees to manage their own leave applications"""

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied.')

        return redirect('core:dashboard')



    employee = request.user.employee



    # Only allow regular employees, not managers/admins/team leaders

    if employee.is_superadmin or employee.is_admin or employee.is_manager or employee.is_team_leader or employee.is_hr:

        # Redirect managers/admins to the full leave management page
        return redirect('accounts:leave_list')


    # Get all employee's leave applications for statistics
    all_leaves = Leave.objects.filter(employee=employee)
    
    # Calculate statistics
    total_leaves = all_leaves.count()
    pending_leaves = all_leaves.filter(status='pending').count()
    approved_leaves = all_leaves.filter(status='approved').count()
    rejected_leaves = all_leaves.filter(status='rejected').count()

    # Get paginated leaves for display
    leaves_qs = all_leaves.order_by('-created_at')
    page = request.GET.get('page', 1)
    paginator = Paginator(leaves_qs, 10)  # 10 items per page for employees

    try:
        leaves = paginator.page(page)

    except PageNotAnInteger:
        leaves = paginator.page(1)

    except EmptyPage:
        leaves = paginator.page(paginator.num_pages)


    context = {
        'leaves': leaves,
        'title': 'My Leave Applications',
        'is_employee_view': True,
        'total_leaves': total_leaves,
        'pending_leaves': pending_leaves,
        'approved_leaves': approved_leaves,
        'rejected_leaves': rejected_leaves,
    }

    return render(request, 'accounts/my_leaves.html', context)



@login_required

@require_http_methods(["GET", "POST"])

def apply_leave(request):

    """Professional view for employees to apply for leave with comprehensive validation and error handling"""

    # Check user permissions

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied. You must be an employee to apply for leave.')

        return redirect('core:dashboard')



    employee = request.user.employee



    # Check if employee is active

    if employee.employment_status != 'active':

        messages.error(request, 'You cannot apply for leave as your employment status is not active.')

        return redirect('core:dashboard')



    if request.method == 'POST':

        form = LeaveApplicationForm(request.POST, employee=employee)



        if form.is_valid():

            try:

                # Use transaction to ensure data consistency

                with transaction.atomic():

                    leave = form.save(commit=False)

                    leave.employee = employee

                    leave.status = 'pending'

                    leave.save()



                    # Create approval request for employee's manager

                    if employee.reports_to:

                        LeaveApproval.objects.create(

                            leave=leave,

                            approver=employee.reports_to,

                            status='pending'

                        )



                        # Log the leave application

                        logger.info(f"Leave application created: {leave.id} by {employee.user.username} for manager approval")



                        messages.success(request,

                            f'Your leave application has been submitted successfully and is pending approval from {employee.reports_to.user.get_full_name()}.')

                    else:

                        # If no manager, auto-approve for admins/superadmins

                        if employee.is_admin or employee.is_superadmin:

                            leave.status = 'approved'

                            leave.approved_by = employee

                            leave.approved_at = timezone.now()

                            leave.save()



                            logger.info(f"Leave application auto-approved: {leave.id} by {employee.user.username} (no manager)")

                            messages.success(request, 'Your leave application has been approved automatically.')

                        else:

                            messages.warning(request,

                                'Your leave application has been submitted but no manager is assigned for approval. Please contact HR.')



                    # Role-based redirect

                    if employee.is_superadmin or employee.is_admin or employee.is_manager or employee.is_team_leader or employee.is_hr:

                        return redirect('accounts:leave_list')

                    else:

                        return redirect('accounts:my_leaves')



            except Exception as e:

                logger.error(f"Error creating leave application for {employee.user.username}: {str(e)}")

                messages.error(request, 'An error occurred while submitting your leave application. Please try again.')

                return redirect('accounts:apply_leave')

        else:

            # Form validation failed

            messages.error(request, 'Please correct the errors below and try again.')

    else:

        form = LeaveApplicationForm(employee=employee)



    # Determine back URL based on user role

    if employee.is_superadmin or employee.is_admin or employee.is_manager or employee.is_team_leader or employee.is_hr:

        back_url = 'accounts:leave_list'

        back_text = 'Back to Leave Management'

    else:

        back_url = 'accounts:my_leaves'

        back_text = 'Back to My Leave'



    context = {

        'form': form,

        'title': 'Apply for Leave',

        'back_url': back_url,

        'back_text': back_text,

        'employee': employee,

    }

    return render(request, 'accounts/apply_leave.html', context)





@login_required

def leave_list(request):

    """View to list leave applications based on user role"""

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied.')

        return redirect('core:dashboard')



    employee = request.user.employee

    form = LeaveFilterForm(request.GET, user=request.user)



    # Base queryset

    leaves = Leave.objects.all()



    # Apply role-based filtering

    if employee.is_superadmin or employee.is_admin or employee.is_hr:

        # Admin/Superadmin/HR can see all leaves

        pass

    elif employee.is_manager:

        # Managers can see leaves of their direct reports and team leaders

        leaves = leaves.filter(

            Q(employee__reports_to=employee) |

            Q(employee__team_leader__reports_to=employee) |

            Q(employee=employee)

        )

    elif employee.is_team_leader:

        # Team leaders can see leaves of their team members only

        leaves = leaves.filter(

            Q(employee__team_leader=employee) |

            Q(employee=employee)

        )

    else:

        # Regular employees can only see their own leaves

        leaves = leaves.filter(employee=employee)



    # Apply form filters

    if form.is_valid():

        if form.cleaned_data.get('status'):

            leaves = leaves.filter(status=form.cleaned_data['status'])

        if form.cleaned_data.get('leave_type'):

            leaves = leaves.filter(leave_type=form.cleaned_data['leave_type'])

        if form.cleaned_data.get('employee'):

            leaves = leaves.filter(employee=form.cleaned_data['employee'])

        if form.cleaned_data.get('date_from'):

            leaves = leaves.filter(start_date__gte=form.cleaned_data['date_from'])

        if form.cleaned_data.get('date_to'):

            leaves = leaves.filter(end_date__lte=form.cleaned_data['date_to'])



    # Get pending approvals for this user

    pending_approvals = LeaveApproval.objects.filter(

        approver=employee,

        status='pending'

    ).select_related('leave__employee')



    # Pagination

    page = request.GET.get('page', 1)

    paginator = Paginator(leaves, 20)  # 20 items per page

    try:

        leaves = paginator.page(page)

    except PageNotAnInteger:

        leaves = paginator.page(1)

    except EmptyPage:

        leaves = paginator.page(paginator.num_pages)



    context = {

        'leaves': leaves,

        'form': form,

        'pending_approvals': pending_approvals,

        'title': 'Leave Management',

    }

    return render(request, 'accounts/leave_list.html', context)





@login_required

def leave_details(request, pk):

    """View to display leave application details"""

    leave = get_object_or_404(Leave, pk=pk)



    # Check permissions

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied.')

        return redirect('core:dashboard')



    employee = request.user.employee

    can_approve = False


    if employee.is_superadmin or employee.is_admin:
        can_approve = True
    elif employee.is_manager:
        # Managers can approve leaves of their direct reports and team leaders
        can_approve = (
            leave.employee.reports_to == employee or
            leave.employee.team_leader and leave.employee.team_leader.reports_to == employee
        )
    elif employee.is_team_leader:
        # Team leaders can approve leaves of their team members
        can_approve = leave.employee.team_leader == employee
    elif employee.is_hr:
        # HR can approve all leaves except superadmins
        can_approve = leave.employee.role != 'superadmin'
        can_reject = True


    # Employees can view their own leaves
    if leave.employee != employee and not can_approve and not can_reject:

        messages.error(request, 'You do not have permission to view this leave application.')

        return redirect('accounts:leave_list')



    context = {

        'leave': leave,

        'can_approve': can_approve,

        'approvals': leave.approvals.all().select_related('approver'),

        'title': f'Leave Details - {leave.employee.full_name}',

    }

    return render(request, 'accounts/leave_details.html', context)





@login_required

@require_http_methods(["POST"])

def approve_leave(request, pk):

    """Professional view to approve a leave application with comprehensive validation and error handling"""

    leave = get_object_or_404(Leave, pk=pk)



    # Check user permissions

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied. You must be logged in as an employee to approve leave.')

        return redirect('core:dashboard')



    employee = request.user.employee



    # Check if user can approve this leave

    can_approve = False

    if employee.is_superadmin or employee.is_admin:

        can_approve = True

    elif employee.is_hr:

        # HR can approve all leaves except superadmins

        can_approve = leave.employee.role != 'superadmin'

    elif employee.is_manager:

        can_approve = (

            leave.employee.reports_to == employee or

            leave.employee.team_leader and leave.employee.team_leader.reports_to == employee

        )

    elif employee.is_team_leader:

        can_approve = leave.employee.team_leader == employee



    if not can_approve:

        messages.error(request, 'You do not have permission to approve this leave application.')

        return redirect('accounts:leave_details', pk=pk)



    # Validate approval comments

    comments = request.POST.get('comments', '').strip()

    if not comments:

        messages.error(request, 'Approval comments are required.')

        return redirect('accounts:leave_details', pk=pk)



    try:

        # Use transaction to ensure data consistency

        with transaction.atomic():

            # Check if leave is still pending (double-check after locking)

            leave.refresh_from_db()

            if leave.status != 'pending':

                messages.error(request, 'This leave application has already been processed.')

                return redirect('accounts:leave_details', pk=pk)



            # Get or create approval record

            approval, created = LeaveApproval.objects.get_or_create(

                leave=leave,

                approver=employee,

                defaults={

                    'status': 'approved',

                    'comments': comments

                }

            )



            if not created:

                if approval.status != 'pending':

                    messages.error(request, 'You have already processed this leave application.')

                    return redirect('accounts:leave_details', pk=pk)

                approval.status = 'approved'

                approval.comments = comments



            approval.approved_at = timezone.now()

            approval.save()



            # Update leave status

            leave.status = 'approved'

            leave.approved_by = employee

            leave.approved_at = timezone.now()

            leave.save()



            # Log the approval

            logger.info(f"Leave application {leave.id} approved by {employee.user.username} for employee {leave.employee.user.username}")



            messages.success(request,

                f'Leave application for {leave.employee.user.get_full_name()} has been approved successfully.')



            # Role-based redirect

            if employee.is_superadmin or employee.is_admin or employee.is_manager or employee.is_team_leader or employee.is_hr:

                return redirect('accounts:leave_list')

            else:

                return redirect('accounts:my_leaves')



    except Exception as e:

        logger.error(f"Error approving leave application {pk} by {employee.user.username}: {str(e)}")

        messages.error(request, 'An error occurred while approving the leave application. Please try again.')

        return redirect('accounts:leave_details', pk=pk)





@login_required

@require_http_methods(["POST"])

def reject_leave(request, pk):

    """Professional view to reject a leave application with comprehensive validation and error handling"""

    leave = get_object_or_404(Leave, pk=pk)



    # Check user permissions

    if not hasattr(request.user, 'employee'):

        messages.error(request, 'Access denied. You must be logged in as an employee to reject leave.')

        return redirect('core:dashboard')



    employee = request.user.employee



    # Check if user can approve this leave

    can_approve = False

    if employee.is_superadmin or employee.is_admin:

        can_approve = True

    elif employee.is_hr:

        # HR can approve all leaves except superadmins

        can_approve = leave.employee.role != 'superadmin'

    elif employee.is_manager:

        can_approve = (

            leave.employee.reports_to == employee or

            leave.employee.team_leader and leave.employee.team_leader.reports_to == employee

        )

    elif employee.is_team_leader:

        can_approve = leave.employee.team_leader == employee



    if not can_approve:

        messages.error(request, 'You do not have permission to reject this leave application.')

        return redirect('accounts:leave_details', pk=pk)



    # Validate rejection comments (required for rejections)

    comments = request.POST.get('comments', '').strip()

    if not comments:

        messages.error(request, 'Rejection comments are required. Please explain why the leave was rejected.')

        return redirect('accounts:leave_details', pk=pk)



    try:

        # Use transaction to ensure data consistency

        with transaction.atomic():

            # Check if leave is still pending (double-check after locking)

            leave.refresh_from_db()

            if leave.status != 'pending':

                messages.error(request, 'This leave application has already been processed.')

                return redirect('accounts:leave_details', pk=pk)



            # Get or create approval record

            approval, created = LeaveApproval.objects.get_or_create(

                leave=leave,

                approver=employee,

                defaults={

                    'status': 'rejected',

                    'comments': comments

                }

            )



            if not created:

                if approval.status != 'pending':

                    messages.error(request, 'You have already processed this leave application.')

                    return redirect('accounts:leave_details', pk=pk)

                approval.status = 'rejected'

                approval.comments = comments



            approval.approved_at = timezone.now()

            approval.save()



            # Update leave status

            leave.status = 'rejected'

            leave.approved_by = employee

            leave.approved_at = timezone.now()

            leave.save()



            # Log the rejection

            logger.info(f"Leave application {leave.id} rejected by {employee.user.username} for employee {leave.employee.user.username}: {comments}")



            messages.success(request,

                f'Leave application for {leave.employee.user.get_full_name()} has been rejected.')



            # Role-based redirect

            if employee.is_superadmin or employee.is_admin or employee.is_manager or employee.is_team_leader or employee.is_hr:

                return redirect('accounts:leave_list')

            else:

                return redirect('accounts:my_leaves')



    except Exception as e:

        logger.error(f"Error rejecting leave application {pk} by {employee.user.username}: {str(e)}")

        messages.error(request, 'An error occurred while rejecting the leave application. Please try again.')


# ==================== MIS (Management Information System) Views ====================

@login_required
@mis_access_required
def mis_dashboard(request):
    """Main MIS Dashboard - Role-based access"""
    
    employee = request.user.employee
    queryset = get_mis_queryset(employee)
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    team_leader_id = request.GET.get('team_leader')
    status = request.GET.get('status')
    
    # Apply filters
    if from_date:
        queryset = queryset.filter(login_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(login_date__lte=to_date)
    if team_leader_id:
        queryset = queryset.filter(team_leader_id=team_leader_id)
    if status:
        queryset = queryset.filter(status=status)
    
    # Calculate aggregates
    total_cases = queryset.count()
    total_login_amount = queryset.aggregate(total=Sum('login_amount'))['total'] or 0
    total_disbursed_amount = queryset.aggregate(total=Sum('disbursed_amount'))['total'] or 0
    
    # Get team leaders for filter (only for managers and admins)
    team_leaders = []
    if employee.is_manager:
        team_leaders = User.objects.filter(
            employee__role='team_leader',
            employee__reports_to=employee
        ).order_by('first_name', 'last_name')
    elif employee.is_admin or employee.is_superadmin:
        team_leaders = User.objects.filter(
            employee__role='team_leader'
        ).order_by('first_name', 'last_name')
    
    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_cases': total_cases,
        'total_login_amount': total_login_amount,
        'total_disbursed_amount': total_disbursed_amount,
        'team_leaders': team_leaders,
        'employee': employee,
        'filter_form': MISFilterForm(user=request.user, initial=request.GET),
    }
    
    if employee.is_team_leader:
        return render(request, 'accounts/mis/team_leader_dashboard.html', context)
    elif employee.is_manager:
        return render(request, 'accounts/mis/manager_dashboard.html', context)
    else:  # Admin/Superadmin
        return render(request, 'accounts/mis/admin_dashboard.html', context)


@login_required
@team_leader_required
def mis_create(request):
    """Create new MIS Report - Team Leader only"""
    
    if request.method == 'POST':
        form = MISReportForm(request.POST, user=request.user)
        if form.is_valid():
            mis_report = form.save(commit=False)
            mis_report.created_by = request.user
            mis_report.team_leader = request.user  # Auto-assign team leader
            
            # Auto-assign manager
            if request.user.employee.manager:
                mis_report.manager = request.user.employee.manager.user
            
            mis_report.save()
            messages.success(request, 'MIS Report created successfully!')
            return redirect('accounts:mis_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = MISReportForm(user=request.user)
    
    return render(request, 'accounts/mis/mis_form.html', {
        'form': form,
        'title': 'Create MIS Report',
        'action': 'Create'
    })


@login_required
@team_leader_required
def mis_edit(request, pk):
    """Edit MIS Report - Team Leader can only edit their own"""
    
    mis_report = get_object_or_404(MISReport, pk=pk, team_leader=request.user)
    
    if request.method == 'POST':
        form = MISReportForm(request.POST, instance=mis_report, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'MIS Report updated successfully!')
            return redirect('accounts:mis_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = MISReportForm(instance=mis_report, user=request.user)
    
    return render(request, 'accounts/mis/mis_form.html', {
        'form': form,
        'mis_report': mis_report,
        'title': 'Edit MIS Report',
        'action': 'Update'
    })


@login_required
@team_leader_required
def mis_delete(request, pk):
    """Delete MIS Report - Team Leader can only delete their own"""
    
    mis_report = get_object_or_404(MISReport, pk=pk, team_leader=request.user)
    
    if request.method == 'POST':
        mis_report.delete()
        messages.success(request, 'MIS Report deleted successfully!')
        return redirect('accounts:mis_dashboard')
    
    return render(request, 'accounts/mis/mis_confirm_delete.html', {
        'mis_report': mis_report
    })


@login_required
@admin_required
def mis_admin_edit(request, pk):
    """Admin can edit any MIS Report"""
    
    mis_report = get_object_or_404(MISReport, pk=pk)
    
    if request.method == 'POST':
        form = MISReportForm(request.POST, instance=mis_report, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'MIS Report updated successfully!')
            return redirect('accounts:mis_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = MISReportForm(instance=mis_report, user=request.user)
    
    return render(request, 'accounts/mis/mis_form.html', {
        'form': form,
        'mis_report': mis_report,
        'title': 'Edit MIS Report',
        'action': 'Update'
    })


@login_required
@manager_required
def mis_manager_edit(request, pk):
    """Manager can edit MIS reports of their team leaders"""
    
    employee = request.user.employee
    queryset = get_mis_queryset(employee)
    mis_report = get_object_or_404(queryset, pk=pk)
    
    if request.method == 'POST':
        form = MISReportForm(request.POST, instance=mis_report, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'MIS Report updated successfully!')
            return redirect('accounts:mis_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = MISReportForm(instance=mis_report, user=request.user)
    
    return render(request, 'accounts/mis/mis_form.html', {
        'form': form,
        'mis_report': mis_report,
        'title': 'Edit MIS Report',
        'action': 'Update'
    })


@login_required
@admin_required
def mis_admin_delete(request, pk):
    """Admin can delete any MIS Report"""
    
    mis_report = get_object_or_404(MISReport, pk=pk)
    
    if request.method == 'POST':
        mis_report.delete()
        messages.success(request, 'MIS Report deleted successfully!')
        return redirect('accounts:mis_dashboard')
    
    return render(request, 'accounts/mis/mis_confirm_delete.html', {
        'mis_report': mis_report
    })


@login_required
@mis_access_required
def mis_export_excel(request):
    """Export MIS Reports to Excel"""
    
    employee = request.user.employee
    queryset = get_mis_queryset(employee)
    
    # Apply same filters as dashboard
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    team_leader_id = request.GET.get('team_leader')
    status = request.GET.get('status')
    
    if from_date:
        queryset = queryset.filter(login_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(login_date__lte=to_date)
    if team_leader_id:
        queryset = queryset.filter(team_leader_id=team_leader_id)
    if status:
        queryset = queryset.filter(status=status)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIS Report"
    
    # Define headers
    headers = [
        'Login Date', 'App ID', 'Customer Name', 'Mobile Number', 'Applicant Type',
        'DOB', 'Age', 'PAN No', 'Salary', 'Company Name', 'Login Amount', 
        'Disbursed Amount', 'Product', 'Tenure (Months)', 'Bank', 'Location',
        'Status', 'Current Status', 'Team Leader', 'Manager', 'Banker Name', 'Banker No'
    ]
    
    # Add headers with styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, mis in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=mis.login_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_num, column=2, value=mis.app_id)
        ws.cell(row=row_num, column=3, value=mis.customer_name)
        ws.cell(row=row_num, column=4, value=mis.mobile_number)
        ws.cell(row=row_num, column=5, value=mis.get_applicant_type_display())
        ws.cell(row=row_num, column=6, value=mis.dob.strftime('%d-%m-%Y'))
        ws.cell(row=row_num, column=7, value=mis.age)
        ws.cell(row=row_num, column=8, value=mis.pan_no)
        ws.cell(row=row_num, column=9, value=mis.salary)
        ws.cell(row=row_num, column=10, value=mis.company_name)
        ws.cell(row=row_num, column=11, value=mis.login_amount)
        ws.cell(row=row_num, column=12, value=mis.disbursed_amount or 0)
        ws.cell(row=row_num, column=13, value=mis.get_product_display())
        ws.cell(row=row_num, column=14, value=mis.tanure)
        ws.cell(row=row_num, column=15, value=mis.bank)
        ws.cell(row=row_num, column=16, value=mis.location)
        ws.cell(row=row_num, column=17, value=mis.get_status_display())
        ws.cell(row=row_num, column=18, value=mis.current_status or '')
        ws.cell(row=row_num, column=19, value=mis.team_leader.get_full_name() if mis.team_leader else '')
        ws.cell(row=row_num, column=20, value=mis.manager.get_full_name() if mis.manager else '')
        ws.cell(row=row_num, column=21, value=mis.banker_name or '')
        ws.cell(row=row_num, column=22, value=mis.banker_no or '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=MIS_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def mis_detail(request, pk):
    """Detailed view of a single MIS report"""
    mis_report = get_object_or_404(MISReport, pk=pk)
    
    # Check permissions
    employee = request.user.employee
    queryset = get_mis_queryset(employee)
    
    if not queryset.filter(pk=pk).exists():
        messages.error(request, "You don't have permission to view this MIS report.")
        return redirect('accounts:mis_dashboard')
    
    context = {
        'mis_report': mis_report,
        'employee': employee,
    }
    
    # Choose template based on user role
    if employee.is_admin or employee.is_superadmin:
        template = 'accounts/mis/admin_detail.html'
    elif employee.is_manager:
        template = 'accounts/mis/manager_detail.html'
    else:  # team_leader
        template = 'accounts/mis/team_leader_detail.html'
    
    return render(request, template, context)


def get_mis_queryset(employee):
    """Get MIS queryset based on user role"""
    
    if employee.is_team_leader:
        return MISReport.objects.filter(team_leader=employee.user)
    elif employee.is_manager:
        return MISReport.objects.filter(team_leader__employee__reports_to=employee)
    elif employee.is_admin or employee.is_superadmin:
        return MISReport.objects.all()
    else:
        return MISReport.objects.none()

        return redirect('accounts:leave_details', pk=pk)
